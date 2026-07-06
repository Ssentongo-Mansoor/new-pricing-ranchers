"""Daily product-level sales import (Month-over-Month product pivot).

Parses the Odoo "Month-over-Month Pivot Invoices Analysis" export and loads it
into ``sales_history`` at customer/product/year/month grain. This feeds the
product reports (products-by-month, velocity, customer product mix).

Designed to run repeatedly from the Admin screen:

* The file is AUTHORITATIVE for the months it contains. We read which
  (year, month) pairs appear in the file, delete existing sales_history rows for
  exactly those months, then insert the file's rows. Months not in the file are
  left untouched. So a daily month-to-date pivot keeps the current month fresh,
  and re-uploading the full history just rebuilds it identically.
* After loading, every sales_history row is re-linked to a catalogue Product
  (token match + imports/product_map.csv overrides), so product reports group
  correctly and off-catalogue items are omitted.

Returns a summary dict.
"""
import os
import re
import csv
import datetime
import collections

from extensions import db
from models import Customer, Product, SalesHistory

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAP_FILE = os.path.join(HERE, "imports", "product_map.csv")

STOP = set("THE A OF KG KGS GR GRS G ML L PC PCS X CATERING RR SALES WHOLE WITH "
           "FOR PER AND".split())


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _nlead(s):
    return len(s) - len(s.lstrip(" "))


def _month(lbl):
    if not lbl:
        return None
    try:
        return datetime.datetime.strptime(str(lbl).strip(), "%B %Y")
    except ValueError:
        return None


def _toks(s):
    s = (s or "").upper()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"\[.*?\]", " ", s)
    s = re.sub(r"[^A-Z ]", " ", s)
    return set(w for w in s.split() if len(w) > 1 and w not in STOP)


def _build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def relink_products():
    """Re-link every sales_history row to a catalogue Product. Returns (linked, total, pct)."""
    prods = [(p.id, p.article_no, p.description) for p in db.session.scalars(db.select(Product))]
    ptoks = [(pid, _toks(desc)) for pid, art, desc in prods]
    by_art = {str(art).strip().upper(): pid for pid, art, _ in prods}

    def best(htk):
        bp, bs = None, 0
        for pid, pt in ptoks:
            if not pt:
                continue
            inter = len(htk & pt)
            if not inter:
                continue
            j = inter / len(htk | pt)
            score = j + (0.5 if (pt <= htk or htk <= pt) else 0)
            if score > bs:
                bs, bp = score, pid
        return bp, bs

    manual = {}
    if os.path.exists(MAP_FILE):
        with open(MAP_FILE, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                nm = (row.get("product_name") or "").strip()
                art = (row.get("article_no") or "").strip().upper()
                if not nm:
                    continue
                manual[nm] = by_art.get(art) if art not in ("", "OMIT", "NONE") else None

    rev_by = collections.defaultdict(float)
    rows = list(db.session.scalars(db.select(SalesHistory)))
    for s in rows:
        rev_by[s.product] += float(s.revenue or 0)

    name_to_pid = {}
    for name in rev_by:
        if name in manual:
            name_to_pid[name] = manual[name]
            continue
        pid, score = best(_toks(name))
        name_to_pid[name] = pid if (pid and score >= 0.5) else None

    n_linked = 0
    for s in rows:
        s.product_id = name_to_pid.get(s.product)
        if s.product_id:
            n_linked += 1
    db.session.commit()

    total = sum(rev_by.values()) or 1.0
    linked_rev = sum(r for nm, r in rev_by.items() if name_to_pid.get(nm))
    return n_linked, len(rows), round(linked_rev / total * 100, 1)


def import_monthly_pivot(path):
    """Load the product pivot; replace only the months present in the file."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    match = _build_matcher()

    parsed = []            # dicts ready for SalesHistory
    months = set()         # (year, month) present in file
    cache = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        grid = list(ws.iter_rows(values_only=True))
        if len(grid) < 4:
            continue
        month_row = grid[1]
        mcols = []         # (untaxed_col_index, year, month)
        for ci, val in enumerate(month_row):
            dt = _month(val)
            if dt:
                mcols.append((ci, dt.year, dt.month))
        cust = None
        for r in grid[3:]:
            a = r[0]
            if a is None:
                continue
            a = str(a)
            lead = _nlead(a)
            name = a.strip()
            if lead == 0 and name == "Total":
                continue
            if lead == 5:
                cust = name
            elif lead == 10 and cust is not None:
                if cust not in cache:
                    cache[cust] = match(cust)
                cid = cache[cust]
                for ci, y, m in mcols:
                    u = r[ci]
                    q = r[ci + 1] if ci + 1 < len(r) else None
                    try:
                        u = float(u)
                    except (TypeError, ValueError):
                        u = 0.0
                    if u == 0:
                        continue
                    try:
                        q = float(q)
                    except (TypeError, ValueError):
                        q = 0.0
                    months.add((y, m))
                    parsed.append(dict(
                        customer_id=cid, customer_name=cust, product=name,
                        year=y, month=m, revenue=round(u, 2), quantity=q,
                        is_return=(u < 0)))

    if not parsed:
        raise ValueError("No product rows found. Is this the Month-over-Month pivot export?")

    # replace only the months present in this file
    month_idx = {y * 12 + m for (y, m) in months}
    for s in list(db.session.scalars(db.select(SalesHistory))):
        if (s.year * 12 + s.month) in month_idx:
            db.session.delete(s)
    db.session.flush()

    for row in parsed:
        db.session.add(SalesHistory(**row))
    db.session.commit()

    linked, total_rows, pct = relink_products()
    mlabels = sorted(months)
    span = (f"{datetime.date(mlabels[0][0], mlabels[0][1], 1):%b %Y}"
            f" – {datetime.date(mlabels[-1][0], mlabels[-1][1], 1):%b %Y}") if mlabels else ""
    return {
        "rows": len(parsed), "months": len(months), "span": span,
        "customers": len(cache),
        "matched_customers": sum(1 for v in cache.values() if v),
        "linked_pct": pct, "total_history_rows": total_rows,
    }
