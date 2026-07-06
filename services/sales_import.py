"""Daily sales / financial data import.

Parses the Odoo Excel exports the finance team produces and upserts them into
the ``invoice`` table that all sales analytics read from. Designed to be run
repeatedly (daily) from the Admin screen:

* Rows are matched to an existing invoice by ``number`` and UPDATED in place;
  new numbers are inserted. Nothing is ever deleted, so a daily file may be
  the full year-to-date export or just the new days -- both work, and
  re-uploading the same file changes nothing.
* Two layouts are supported (chosen by the user on the upload screen):
    - "invoices"     -> "Invoice List Customer-Wise" export (positive amounts)
    - "credit_notes" -> "Credit Notes Customer List"   export (negative amounts)

Returns a summary dict: rows read, inserted, updated, customers matched.
"""
import re
import collections
import datetime

from openpyxl import load_workbook

from extensions import db
from models import Customer, Invoice


# Column maps (0-based) for each export layout.
LAYOUTS = {
    "invoices": {
        "number": 0, "customer": 2, "date": 4, "due": 5, "salesperson": 9,
        "untaxed": 11, "total": 12, "currency": 13, "payment": 14,
        "company_type": 7, "efris": 1,
    },
    "credit_notes": {
        "number": 0, "customer": 1, "date": 3, "due": 4, "salesperson": None,
        "untaxed": 6, "total": 7, "currency": None, "payment": 9,
        "company_type": None, "efris": 11,
    },
}


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _num(v):
    if isinstance(v, str):
        v = v.replace(",", "").strip()
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def _pick_sheet(wb):
    if "Sheet1" in wb.sheetnames:
        return wb["Sheet1"]
    return wb[wb.sheetnames[0]]


def import_file(path, layout="invoices"):
    """Upsert one Odoo export into the invoice table. Returns a summary dict."""
    if layout not in LAYOUTS:
        raise ValueError(f"Unknown layout '{layout}'")
    cmap = LAYOUTS[layout]
    wb = load_workbook(path, data_only=True)
    ws = _pick_sheet(wb)
    match = _build_matcher()

    def cell(row, key):
        idx = cmap[key]
        return row[idx] if idx is not None and idx < len(row) else None

    existing = {inv.number: inv for inv in db.session.scalars(db.select(Invoice))}
    cust_cache = {}
    read = inserted = updated = matched = 0

    for r in ws.iter_rows(min_row=2, values_only=True):
        num = cell(r, "number")
        if num is None:
            continue
        num = str(num).strip()
        if not num:
            continue
        # skip year group-header rows like "2024 (123)"
        if "(" in num and num[:1].isdigit():
            continue
        read += 1

        cname = (cell(r, "customer") or "")
        cname = cname.strip() if isinstance(cname, str) else str(cname)
        if cname not in cust_cache:
            cust_cache[cname] = match(cname)
        cid = cust_cache[cname]
        if cid:
            matched += 1

        sp = cell(r, "salesperson")
        sp = (sp or "").strip() if isinstance(sp, str) else None
        if sp:
            sp = re.sub(r"\s*\(.*?\)\s*$", "", sp) or None

        currency = cell(r, "currency")
        currency = (currency or "UGX").strip() if isinstance(currency, str) else "UGX"
        payment = cell(r, "payment")
        payment = (payment or "").strip() if isinstance(payment, str) else None
        ctype = cell(r, "company_type")
        ctype = (ctype or "").strip() if isinstance(ctype, str) else None
        efris = cell(r, "efris")
        efris = str(efris) if efris not in (None, "") else None

        fields = dict(
            customer_id=cid, customer_name=cname, salesperson=sp,
            invoice_date=_date(cell(r, "date")), due_date=_date(cell(r, "due")),
            untaxed=_num(cell(r, "untaxed")), total=_num(cell(r, "total")),
            currency=currency or "UGX", payment_status=payment or None,
            company_type=ctype or None, efris=efris)

        inv = existing.get(num)
        if inv:
            for k, v in fields.items():
                setattr(inv, k, v)
            updated += 1
        else:
            inv = Invoice(number=num, **fields)
            db.session.add(inv)
            existing[num] = inv
            inserted += 1

        if read % 5000 == 0:
            db.session.flush()

    db.session.commit()
    return {
        "layout": layout, "read": read, "inserted": inserted,
        "updated": updated, "customers": len(cust_cache),
        "matched_customers": sum(1 for v in cust_cache.values() if v),
        "matched_rows": matched,
    }
