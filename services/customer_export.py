"""Build a customisable Excel export of the customer list.

Callers choose the columns and pass an already-filtered list of customers plus
the set of active-customer ids. Revenue/last-order figures come from invoice
history plus live orders.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from extensions import db
from models import Invoice, SalesOrder
from services import targets as tsvc

# key -> (header, width)
COLUMNS = [
    ("name", "Customer", 34),
    ("status", "Status", 12),
    ("segment", "Type", 12),
    ("category", "Category", 20),
    ("reps", "Rep(s)", 24),
    ("phone", "Phone", 18),
    ("email", "Email", 26),
    ("tax_id", "Tax ID", 14),
    ("address", "Address / City", 26),
    ("area", "Area", 18),
    ("procurement", "Procurement contact", 34),
    ("chef", "Chef contact", 34),
    ("other", "Other contact", 34),
    ("delivery_days", "Delivery days", 20),
    ("delivery_window", "Delivery window", 16),
    ("delivery_notes", "Delivery notes", 26),
    ("payment_terms", "Payment terms", 18),
    ("account_status", "Account status", 14),
    ("last_order", "Last order", 12),
    ("recent_rev", "Recent 3-mo (UGX)", 18),
    ("lifetime_rev", "Lifetime (UGX)", 18),
    ("created", "Added", 12),
]
COL_LABELS = {k: h for k, h, _ in COLUMNS}
DEFAULT_COLS = ["name", "status", "segment", "category", "reps", "phone", "email",
                "tax_id", "last_order", "recent_rev", "lifetime_rev"]


def _idx_to_month(idx):
    return date((idx - 1) // 12, (idx - 1) % 12 + 1, 1)


def revenue_maps():
    """Return per-customer {last_idx, recent, lifetime}."""
    today = date.today()
    y, m = today.year, today.month - 2
    while m <= 0:
        m += 12
        y -= 1
    recent_cut = date(y, m, 1)
    last = {}
    recent = {}
    lifetime = {}
    for i in db.session.scalars(db.select(Invoice).where(
            Invoice.customer_id.isnot(None), Invoice.currency == "UGX",
            Invoice.payment_status != "Reversed", Invoice.invoice_date.isnot(None))):
        cid = i.customer_id
        v = float(i.untaxed or 0)
        lifetime[cid] = lifetime.get(cid, 0.0) + v
        idx = i.invoice_date.year * 12 + i.invoice_date.month
        if v > 0:
            last[cid] = max(last.get(cid, 0), idx)
        if i.invoice_date >= recent_cut:
            recent[cid] = recent.get(cid, 0.0) + v
    for o in db.session.scalars(db.select(SalesOrder).where(
            SalesOrder.customer_id.isnot(None),
            SalesOrder.status.in_(tsvc.CONFIRMED),
            SalesOrder.order_date.isnot(None))):
        cid = o.customer_id
        v = tsvc._ugx(o)
        lifetime[cid] = lifetime.get(cid, 0.0) + v
        idx = o.order_date.year * 12 + o.order_date.month
        last[cid] = max(last.get(cid, 0), idx)
        if o.order_date >= recent_cut:
            recent[cid] = recent.get(cid, 0.0) + v
    return {"last": last, "recent": recent, "lifetime": lifetime}


def _contact(nm, tel, em):
    parts = [p for p in (nm, tel, em) if p]
    return " · ".join(parts)


def _value(col, c, active_ids, maps):
    if col == "name":
        return c.name
    if col == "status":
        return "Active" if c.id in active_ids else "Not active"
    if col == "segment":
        return "Distributor" if (c.segment or "customer") == "distributor" else "Customer"
    if col == "category":
        return c.category.name if c.category else ""
    if col == "reps":
        return ", ".join(r.full_name for r in c.reps)
    if col == "phone":
        return c.phone or ""
    if col == "email":
        return c.email or ""
    if col == "tax_id":
        return c.tax_id or ""
    if col == "address":
        return c.address or ""
    if col == "area":
        return c.area or ""
    if col == "procurement":
        return _contact(c.procurement_name, c.procurement_phone, c.procurement_email)
    if col == "chef":
        return _contact(c.chef_name, c.chef_phone, c.chef_email)
    if col == "other":
        return _contact(c.other_contact_name, c.other_contact_phone, c.other_contact_email)
    if col == "delivery_days":
        return (c.delivery_days or "").replace(",", ", ")
    if col == "delivery_window":
        if c.delivery_time_from or c.delivery_time_to:
            return f"{c.delivery_time_from or '…'}-{c.delivery_time_to or '…'}"
        return ""
    if col == "delivery_notes":
        return c.delivery_notes or ""
    if col == "payment_terms":
        return c.payment_terms or ""
    if col == "account_status":
        return c.account_status or "ok"
    if col == "last_order":
        idx = maps["last"].get(c.id)
        return _idx_to_month(idx).strftime("%b %Y") if idx else ""
    if col == "recent_rev":
        return round(maps["recent"].get(c.id, 0.0))
    if col == "lifetime_rev":
        return round(maps["lifetime"].get(c.id, 0.0))
    if col == "created":
        return c.created_at.strftime("%d %b %Y") if c.created_at else ""
    return ""


def build_workbook(customers, cols, active_ids, title="Customers"):
    cols = [c for c in cols if c in COL_LABELS] or DEFAULT_COLS
    maps = revenue_maps() if any(c in ("last_order", "recent_rev", "lifetime_rev")
                                 for c in cols) else {"last": {}, "recent": {}, "lifetime": {}}
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    head_fill = PatternFill("solid", fgColor="6E7149")
    head_font = Font(bold=True, color="FFFFFF")
    widths = {k: w for k, _h, w in COLUMNS}
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=COL_LABELS[col])
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)
    for ri, c in enumerate(customers, start=2):
        for ci, col in enumerate(cols, start=1):
            v = _value(col, c, active_ids, maps)
            cell = ws.cell(row=ri, column=ci, value=v)
            if col in ("recent_rev", "lifetime_rev"):
                cell.number_format = "#,##0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
