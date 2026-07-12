"""Branded Excel and PDF exports for pricelists, customer pricelists and offers.

Every export leads with the Ranchers Finest logo on a brand-coloured band, the
list/offer name, currency, the exchange rate used (for USD), and effective /
validity dates, styled in the brand palette.
"""
import io
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage,
)

from services.pricing import format_money, effective_line_price
from services import settings, currency as cx

# Brand palette
RF_ORANGE = colors.HexColor("#F47A21")
RF_RED = colors.HexColor("#ED1C24")
RF_AMBER = colors.HexColor("#F9A03F")
RF_INK = colors.HexColor("#0E0E0E")
RF_CHARCOAL = colors.HexColor("#2B2B2B")
RF_CREAM = colors.HexColor("#FDF1E5")

_HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGO_PNG = os.path.join(_HERE, "static", "img", "ranchers-logo.png")


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _vat_off_label(doc):
    """Label for a VAT-off document. "Zero-rated (export)" is the correct URA
    wording only for a genuine export; a local VAT-off list/order is simply VAT
    not applicable (L10)."""
    if (getattr(doc, "market", None) or "local") == "export":
        return "Zero-rated (export)"
    return "VAT not applicable"


def _rate_note(pricelist_or_offer, ccy):
    """Build a human note about the exchange rate used for a USD export."""
    if ccy != "USD":
        return None
    # Use the list's own effective date so an old list shows the rate that
    # applied then, not today's (L5).
    on = (getattr(pricelist_or_offer, "effective_date", None)
          or getattr(pricelist_or_offer, "valid_from", None))
    rate = cx.get_rate("USD", on=on)
    if rate is None:
        return "USD rate: none on file"
    window = f"from {rate.effective_date}"
    if rate.expiry_date:
        window += f" to {rate.expiry_date}"
    return f"Rate used: UGX {float(rate.rate):,.0f} / USD ({window})"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _pdf_header(elements, styles, title, subtitle_lines, width=247 * mm):
    band = Table([[""]], colWidths=[width], rowHeights=[2 * mm])
    band.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), RF_ORANGE)]))

    logo_cell = ""
    if os.path.exists(LOGO_PNG):
        img = RLImage(LOGO_PNG, width=26 * mm, height=18 * mm)
        logo_cell = img

    title_style = ParagraphStyle("rf_title", parent=styles["Title"],
                                 textColor=colors.white, fontSize=18, alignment=0,
                                 leading=22, spaceAfter=0)
    sub_style = ParagraphStyle("rf_sub", parent=styles["Normal"],
                               textColor=colors.white, fontSize=9, leading=12)
    company = settings.get("company_name", "Ranchers Finest U Ltd")
    text_block = [Paragraph(company, title_style),
                  Paragraph(title, sub_style)]
    for ln in subtitle_lines:
        if ln:
            text_block.append(Paragraph(ln, sub_style))

    head = Table([[logo_cell, text_block]], colWidths=[32 * mm, width - 32 * mm])
    head.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), RF_INK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(head)
    elements.append(band)
    elements.append(Spacer(1, 6 * mm))


def _validity_lines(obj):
    lines = []
    eff = getattr(obj, "effective_date", None) or getattr(obj, "valid_from", None)
    vu = getattr(obj, "valid_until", None)
    if eff:
        lines.append(f"Effective: {eff}")
    if vu:
        lines.append(f"Valid until: {vu}")
    return lines


def pricelist_to_pdf(pricelist):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=10 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9)
    sect = ParagraphStyle("sect", parent=styles["Normal"], fontSize=8,
                          textColor=colors.white, leading=10, fontName="Helvetica-Bold")
    elements = []

    vat = _vat_off_label(pricelist) if pricelist.is_zero_rated else f"VAT {pricelist.vat_rate:.0f}% applicable"
    subt = [f"{pricelist.name}",
            f"Currency: {pricelist.currency} · {vat}"]
    if pricelist.price_basis == "kg":
        subt.append("Prices are per kilogram (kg)")
    subt += _validity_lines(pricelist)
    rn = _rate_note(pricelist, pricelist.currency)
    if rn:
        subt.append(rn)
    _pdf_header(elements, styles, "PRICELIST", subt)

    tiers = pricelist.tiers
    has_box = any(l.box_small or l.box_medium or l.box_large for l in pricelist.lines)
    header = ["Art No", "Description", "Pack size"]
    header += [t.label for t in tiers]
    if has_box:
        header += ["Box S", "Box M", "Box L"]
    data = [header]

    # group rows by section
    current = None
    col_count = len(header)
    for line in pricelist.lines:
        if line.section and line.section != current:
            current = line.section
            data.append([Paragraph(current, sect)] + [""] * (col_count - 1))
        row = [line.product.article_no,
               Paragraph(line.product.description, cell),
               line.pack_size or line.product.pack_size or ""]
        for t in tiers:
            eff = effective_line_price(line, t.key)
            txt = format_money(eff["amount"], eff["currency"]) if eff["amount"] is not None else ""
            if eff["is_fixed"]:
                txt += " [fixed]"
            row.append(txt)
        if has_box:
            row += [_num(line.box_small), _num(line.box_medium), _num(line.box_large)]
        data.append(row)

    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), RF_INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, RF_CREAM]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E0D5C8")),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    # shade section rows
    for i, row in enumerate(data):
        if isinstance(row[0], Paragraph) and getattr(row[0], "style", None) is sect:
            style.append(("BACKGROUND", (0, i), (-1, i), RF_ORANGE))
            style.append(("SPAN", (0, i), (-1, i)))
    table.setStyle(TableStyle(style))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf.read()


def offer_to_pdf(offer):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=10 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8.5, leading=11)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, leading=9.5)
    hdr = ParagraphStyle("hdr", parent=styles["Normal"], fontSize=7.8, leading=9,
                         textColor=colors.white, fontName="Helvetica-Bold")
    elements = []

    vat = _vat_off_label(offer) if not offer.vat_applicable else f"VAT {offer.vat_rate:.0f}%"
    src = offer.source_pricelist
    basis = src.price_basis if src is not None else None   # 'kg' | 'pack' | None
    subt = [f"Offer {offer.number} for {offer.customer.name}",
            f"Currency: {offer.currency} · {vat}"]
    if basis == "kg":
        subt.append("Prices are per kilogram (kg)")
    elif basis is None and src is not None:
        subt.append("Prices per kg or per pack — see the Pack column")
    subt += _validity_lines(offer)
    if offer.currency == "USD" and offer.exchange_rate_value:
        subt.append(f"Rate used (stamped): UGX {float(offer.exchange_rate_value):,.0f} / USD")
    _pdf_header(elements, styles, "QUOTATION / QUOTE", subt, width=doc.width)

    qty_hdr = "Qty (kg)" if basis == "kg" else "Qty"
    header = [Paragraph(h, hdr) for h in
              ["Art No", "Description", "Pack size", "Tier", qty_hdr,
               "Unit excl VAT", "Unit incl VAT", "Disc %", "Total incl VAT"]]
    data = [header]
    for l in offer.lines:
        note = " [fixed]" if l.is_fixed else ""
        vf = (1 + offer.vat_rate / 100.0) if (offer.vat_applicable and l.is_vatable) else 1.0
        unit_incl = float(l.unit_price or 0) * vf
        line_incl = float(l.line_total or 0) * vf
        pk = l.pack_size or ""
        if src is not None and basis != "pack":   # annotate wholesale (kg) / mixed; retail is clear
            bl = src.basis_for(getattr(l, "product", None))
            pk = (pk + " · " if pk else "") + ("/kg" if bl == "kg" else "/pack")
        data.append([
            l.article_no or "",
            Paragraph((l.description or "") + note, cell),
            Paragraph(pk, small),
            Paragraph(l.tier_label or "", small),
            _num(l.quantity),
            format_money(l.unit_price, offer.currency),
            format_money(unit_incl, offer.currency),
            f"{l.discount_pct:.0f}" if l.discount_pct else "0",
            format_money(line_incl, offer.currency),
        ])
    table = Table(data, repeatRows=1,
                  colWidths=[15*mm, 37*mm, 12*mm, 19*mm, 9*mm, 21*mm, 21*mm, 11*mm, 22*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), RF_INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, RF_CREAM]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E0D5C8")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 4 * mm))

    totals = [["Subtotal (excl VAT)", format_money(offer.subtotal, offer.currency)]]
    if offer.vat_applicable:
        totals.append([f"VAT ({offer.vat_rate:.0f}%)", format_money(offer.vat_amount, offer.currency)])
    else:
        totals.append(["VAT", _vat_off_label(offer)])
    totals.append(["TOTAL (incl VAT)", format_money(offer.total, offer.currency)])
    tt = Table(totals, colWidths=[40 * mm, 40 * mm], hAlign="RIGHT")
    tt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.6, RF_INK),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, -1), (-1, -1), RF_RED),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tt)
    if offer.notes:
        elements.append(Spacer(1, 5 * mm))
        elements.append(Paragraph(f"<b>Notes:</b> {offer.notes}", cell))
    doc.build(elements)
    buf.seek(0)
    return buf.read()


def _doc_header(elements, styles, title, ref, status=None):
    """Branded header band: logo left, company + document title + reference right."""
    company = settings.get("company_name", "Ranchers Finest U Ltd")
    logo_cell = ""
    if os.path.exists(LOGO_PNG):
        logo_cell = RLImage(LOGO_PNG, width=28 * mm, height=19 * mm)
    title_style = ParagraphStyle("dh_company", parent=styles["Title"],
                                 textColor=colors.white, fontSize=17, alignment=2,
                                 leading=20, spaceAfter=0)
    doc_style = ParagraphStyle("dh_doc", parent=styles["Normal"],
                               textColor=colors.white, fontSize=11, alignment=2, leading=14)
    ref_style = ParagraphStyle("dh_ref", parent=styles["Normal"],
                               textColor=colors.HexColor("#FFE7D2"), fontSize=9, alignment=2)
    right = [Paragraph(company, title_style),
             Paragraph(f"{title}", doc_style),
             Paragraph(ref + (f"  ·  {status}" if status else ""), ref_style)]
    head = Table([[logo_cell, right]], colWidths=[40 * mm, 141 * mm])
    head.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), RF_INK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(head)
    band = Table([[""]], colWidths=[181 * mm], rowHeights=[2.4 * mm])
    band.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), RF_ORANGE)]))
    elements.append(band)
    elements.append(Spacer(1, 6 * mm))


def _info_box(title, rows, styles, value_color=RF_CHARCOAL):
    """A titled panel of label/value rows."""
    lbl = ParagraphStyle("ib_t", parent=styles["Normal"], fontSize=8,
                         textColor=colors.white, fontName="Helvetica-Bold")
    k = ParagraphStyle("ib_k", parent=styles["Normal"], fontSize=8.5,
                       textColor=colors.HexColor("#7a7068"))
    v = ParagraphStyle("ib_v", parent=styles["Normal"], fontSize=9.5,
                       textColor=value_color, leading=12)
    data = [[Paragraph(title, lbl), ""]]
    for key, val in rows:
        if val in (None, ""):
            continue
        data.append([Paragraph(str(key), k), Paragraph(str(val), v)])
    t = Table(data, colWidths=[26 * mm, 62 * mm])
    style = [
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), RF_INK),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#E0D5C8")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.4, colors.HexColor("#E0D5C8")),
    ]
    t.setStyle(TableStyle(style))
    return t


def order_to_pdf(order):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"Order {order.number}",
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=10 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8.5, leading=11)
    elements = []

    _doc_header(elements, styles, "SALES ORDER", order.number, order.status.title())

    c = order.customer
    cust_rows = [("Customer", f"<b>{c.name}</b>"),
                 ("Contact", c.contact_name), ("Phone", c.phone), ("Email", c.email),
                 ("Deliver to", order.delivery_address)]
    vat = _vat_off_label(order) if not order.vat_applicable else f"VAT {order.vat_rate:.0f}% applicable"
    ord_rows = [("Order no", f"<b>{order.number}</b>"),
                ("Status", order.status.replace('_', ' ').title()),
                ("Order date", order.order_date),
                ("Delivery", order.delivery_date),
                ("Customer PO", order.customer_po),
                ("Terms", order.payment_terms),
                ("Currency", order.currency),
                ("VAT", vat)]
    if order.currency == "USD" and order.exchange_rate_value:
        ord_rows.append(("Rate", f"UGX {float(order.exchange_rate_value):,.0f} / USD (stamped)"))

    info = Table([[_info_box("CUSTOMER", cust_rows, styles),
                   _info_box("ORDER DETAILS", ord_rows, styles)]],
                 colWidths=[90 * mm, 91 * mm])
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (0, 0), (0, 0), 0),
                              ("RIGHTPADDING", (1, 0), (1, 0), 0)]))
    elements.append(info)
    elements.append(Spacer(1, 6 * mm))

    show_delivered = order.status in ("in_fulfillment", "pending", "fulfilled",
                                      "ready_for_dispatch", "out_for_delivery", "delivered")
    _osrc = getattr(order, "source_pricelist", None)
    _kg = " (kg)" if (_osrc is not None and _osrc.price_basis == "kg") else ""
    if show_delivered:
        header = ["Art No", "Description", "Pack size", "Ordered" + _kg, "Delivered" + _kg,
                  "Unit Price", "Disc %", "Line Total", "Status"]
        colw = [14*mm, 47*mm, 15*mm, 14*mm, 14*mm, 22*mm, 11*mm, 23*mm, 20*mm]
        status_col = 8
        num_first = 3
    else:
        header = ["Art No", "Description", "Pack size", "Qty" + _kg, "Unit Price",
                  "Disc %", "Line Total", "Status"]
        colw = [16*mm, 58*mm, 18*mm, 12*mm, 24*mm, 12*mm, 25*mm, 16*mm]
        status_col = 7
        num_first = 3
    data = [header]
    for l in order.lines:
        avail = getattr(l, "availability", "available") or "available"
        badge = {"available": "", "out_of_stock": "OUT OF STOCK",
                 "not_delivered": "NOT DELIVERED"}.get(avail, "")
        note = " [fixed]" if l.is_fixed else ""
        if order.vat_applicable and getattr(l, "is_vatable", False):
            note += " *"
        qty_cells = [_num(l.quantity), _num(l.delivered_qty)] if show_delivered else [_num(l.quantity)]
        data.append([
            l.article_no or "", Paragraph((l.description or "") + note, cell),
            l.pack_size or "", *qty_cells,
            format_money(l.unit_price, order.currency),
            f"{l.discount_pct:.0f}" if l.discount_pct else "0",
            format_money(l.line_total, order.currency), badge])
    table = Table(data, repeatRows=1, colWidths=colw)
    tstyle = [
        ("BACKGROUND", (0, 0), (-1, 0), RF_INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.3),
        ("ALIGN", (num_first, 1), (status_col - 1, -1), "RIGHT"),
        ("ALIGN", (status_col, 0), (status_col, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, RF_CREAM]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E0D5C8")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for i, l in enumerate(order.lines, start=1):
        avail = getattr(l, "availability", "available") or "available"
        if avail == "out_of_stock":
            tstyle.append(("TEXTCOLOR", (status_col, i), (status_col, i), RF_AMBER))
            tstyle.append(("FONTNAME", (status_col, i), (status_col, i), "Helvetica-Bold"))
        elif avail == "not_delivered":
            tstyle.append(("TEXTCOLOR", (status_col, i), (status_col, i), RF_RED))
            tstyle.append(("FONTNAME", (status_col, i), (status_col, i), "Helvetica-Bold"))
    table.setStyle(TableStyle(tstyle))
    elements.append(table)
    elements.append(Spacer(1, 4 * mm))

    totals = [["Subtotal (net)", format_money(order.subtotal, order.currency)]]
    if not order.vat_applicable:
        totals.append(["VAT", _vat_off_label(order)])
    elif order.vat_amount:
        totals.append([f"VAT {order.vat_rate:.0f}% (items *)", format_money(order.vat_amount, order.currency)])
    else:
        totals.append(["VAT", "None (no VAT items)"])
    totals.append(["TOTAL", format_money(order.total, order.currency)])
    tt = Table(totals, colWidths=[40 * mm, 40 * mm], hAlign="RIGHT")
    tt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.6, RF_INK),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, -1), (-1, -1), RF_RED),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tt)
    if order.vat_applicable and order.vat_amount:
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph(
            "<font size=7>* VAT charged on these items only. Other items are VAT-exempt.</font>", cell))
    if order.notes:
        elements.append(Spacer(1, 5 * mm))
        elements.append(Paragraph(f"<b>Notes:</b> {order.notes}", cell))
    doc.build(elements)
    buf.seek(0)
    return buf.read()


def delivery_note_to_pdf(order):
    """A delivery note: delivered quantities and a signature block for the customer."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            title=f"Delivery Note {order.dnote_number or order.number}",
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=10 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8.5, leading=11)
    elements = []
    _doc_header(elements, styles, "DELIVERY NOTE",
                order.dnote_number or order.number, None)

    c = order.customer
    cust_rows = [("Customer", f"<b>{c.name}</b>"), ("Contact", c.contact_name),
                 ("Phone", c.phone), ("Deliver to", order.delivery_address)]
    _dnd = order.dnote_at or order.fulfilled_at or order.order_date
    dn_date = _dnd.strftime("%d %b %Y %H:%M") if hasattr(_dnd, "strftime") and hasattr(_dnd, "hour") \
        else (_dnd.strftime("%d %b %Y") if hasattr(_dnd, "strftime") else str(_dnd or "—"))
    dn_rows = [("Delivery note", f"<b>{order.dnote_number or '—'}</b>"),
               ("Order no", order.number),
               ("Date", dn_date),
               ("Customer PO", order.customer_po),
               ("Driver", order.assigned_driver.full_name if order.assigned_driver else "—")]
    info = Table([[_info_box("CUSTOMER", cust_rows, styles),
                   _info_box("DELIVERY", dn_rows, styles)]],
                 colWidths=[90 * mm, 91 * mm])
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (0, 0), (0, 0), 0),
                              ("RIGHTPADDING", (1, 0), (1, 0), 0)]))
    elements.append(info)
    elements.append(Spacer(1, 6 * mm))

    header = ["Art No", "Description", "Pack", "Ordered", "Delivered", "VAT", "Status"]
    colw = [16 * mm, 70 * mm, 18 * mm, 18 * mm, 18 * mm, 12 * mm, 29 * mm]
    data = [header]
    for l in order.lines:
        avail = getattr(l, "availability", "available") or "available"
        badge = {"available": "", "out_of_stock": "OUT OF STOCK",
                 "not_delivered": "NOT DELIVERED"}.get(avail, "")
        vmark = "VAT" if (order.vat_applicable and getattr(l, "is_vatable", False)) else "—"
        data.append([l.article_no or "", Paragraph(l.description or "", cell),
                     l.pack_size or "", _num(l.quantity), _num(l.delivered_qty), vmark, badge])
    table = Table(data, repeatRows=1, colWidths=colw)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), RF_INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (3, 1), (4, -1), "RIGHT"),
        ("ALIGN", (5, 0), (6, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, RF_CREAM]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E0D5C8")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 5 * mm))

    # Value summary (VAT only on the items marked VAT)
    totals = [["Subtotal (net)", format_money(order.subtotal, order.currency)]]
    if not order.vat_applicable:
        totals.append(["VAT", _vat_off_label(order)])
    elif order.vat_amount:
        totals.append([f"VAT {order.vat_rate:.0f}% (VAT items)", format_money(order.vat_amount, order.currency)])
    else:
        totals.append(["VAT", "None (no VAT items)"])
    totals.append(["TOTAL", format_money(order.total, order.currency)])
    tt = Table(totals, colWidths=[55 * mm, 40 * mm], hAlign="RIGHT")
    tt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.6, RF_INK),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, -1), (-1, -1), RF_RED),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tt)
    elements.append(Spacer(1, 10 * mm))
    sign = Table([["Received by (name):", "Signature:", "Date:"]],
                 colWidths=[70 * mm, 60 * mm, 51 * mm])
    sign.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, RF_INK),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
    ]))
    elements.append(Paragraph("Goods received in good order and condition:", cell))
    elements.append(Spacer(1, 4 * mm))
    elements.append(sign)
    doc.build(elements)
    buf.seek(0)
    return buf.read()


def order_to_excel(order):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Order {order.number}"[:31]
    # Delivered column added so "Delivered x Unit Price" reconciles to Line Total
    # (Line Total is computed on the delivered quantity, not the ordered one) (M14).
    headers = ["Art No", "Description", "Pack", "Tier", "Ordered", "Delivered",
               "Unit Price", "Disc %", "Line Total"]
    ncols = len(headers)
    # column indices of the two summary columns (under Disc % / Line Total)
    lbl_col = ncols - 1
    val_col = ncols
    vat = _vat_off_label(order) if not order.vat_applicable else f"VAT {order.vat_rate:.0f}%"
    subt = [f"Order {order.number} for {order.customer.name} · {order.status.title()}",
            f"Currency: {order.currency} · {vat}", f"Order date: {order.order_date}"]
    if order.delivery_date:
        subt.append(f"Requested delivery: {order.delivery_date}")
    if order.customer_po:
        subt.append(f"Customer PO: {order.customer_po}")
    if order.payment_terms:
        subt.append(f"Payment terms: {order.payment_terms}")
    if order.currency == "USD" and order.exchange_rate_value:
        subt.append(f"Rate used (stamped): UGX {float(order.exchange_rate_value):,.0f} / USD")
    row = _xl_header(ws, ncols, "SALES ORDER", subt)

    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_table_header(ws, row, ncols)
    row += 1
    decimals = cx.rounding_for(order.currency)
    fmt = "#,##0" if decimals == 0 else "#,##0." + "0" * decimals
    for l in order.lines:
        note = " [fixed]" if l.is_fixed else ""
        ws.cell(row=row, column=1, value=l.article_no or "")
        ws.cell(row=row, column=2, value=(l.description or "") + note)
        ws.cell(row=row, column=3, value=l.pack_size or "")
        ws.cell(row=row, column=4, value=l.tier_label or "")
        ws.cell(row=row, column=5, value=l.quantity)                     # ordered
        ws.cell(row=row, column=6, value=l.delivered_qty)               # delivered (drives Line Total)
        c7 = ws.cell(row=row, column=7, value=float(l.unit_price or 0)); c7.number_format = fmt
        ws.cell(row=row, column=8, value=l.discount_pct or 0)
        c9 = ws.cell(row=row, column=9, value=float(l.line_total or 0)); c9.number_format = fmt
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).border = _BORDER
        row += 1
    row += 1
    ws.cell(row=row, column=lbl_col, value="Subtotal").font = Font(bold=True)
    ws.cell(row=row, column=val_col, value=float(order.subtotal)).number_format = fmt
    row += 1
    ws.cell(row=row, column=lbl_col,
            value=f"VAT ({order.vat_rate:.0f}%)" if order.vat_applicable else "VAT").font = Font(bold=True)
    if order.vat_applicable:
        ws.cell(row=row, column=val_col, value=float(order.vat_amount)).number_format = fmt
    else:
        # numeric amount cell gets 0 (a number); the label carries the wording
        ws.cell(row=row, column=val_col, value=0).number_format = fmt
        ws.cell(row=row, column=lbl_col, value=f"VAT — {_vat_off_label(order)}").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=lbl_col, value="TOTAL").font = Font(bold=True, color="ED1C24")
    tc = ws.cell(row=row, column=val_col, value=float(order.total)); tc.number_format = fmt
    tc.font = Font(bold=True, color="ED1C24")
    _autosize(ws, ncols)
    return _wb_bytes(wb)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
_THIN = Side(style="thin", color="E0D5C8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _xl_header(ws, ncols, title, subtitle_lines):
    last = get_column_letter(max(ncols, 4))
    if os.path.exists(LOGO_PNG):
        try:
            img = XLImage(LOGO_PNG)
            img.height = 60
            img.width = 86
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.merge_cells(f"B1:{last}1")
    c = ws["B1"]
    c.value = settings.get("company_name", "Ranchers Finest U Ltd")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.alignment = Alignment(vertical="center", horizontal="left")
    for col in range(1, max(ncols, 4) + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="0E0E0E")
    ws.row_dimensions[1].height = 46

    r = 2
    ws.merge_cells(f"A{r}:{last}{r}")
    t = ws[f"A{r}"]
    t.value = title
    t.font = Font(bold=True, size=12, color="F47A21")
    r += 1
    for ln in subtitle_lines:
        if not ln:
            continue
        ws.merge_cells(f"A{r}:{last}{r}")
        ws[f"A{r}"].value = ln
        ws[f"A{r}"].font = Font(size=9, color="2B2B2B")
        r += 1
    return r + 1  # next free row


def _style_table_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor="0E0E0E")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER


def pricelist_to_excel(pricelist):
    wb = Workbook()
    ws = wb.active
    ws.title = (pricelist.name or "Pricelist")[:31]

    tiers = pricelist.tiers
    has_box = any(l.box_small or l.box_medium or l.box_large for l in pricelist.lines)
    headers = ["Art No", "Barcode", "Description", "Pack size"] + [t.label for t in tiers]
    if has_box:
        headers += ["Box S", "Box M", "Box L"]
    ncols = len(headers)

    vat = _vat_off_label(pricelist) if pricelist.is_zero_rated else f"VAT {pricelist.vat_rate:.0f}% applicable"
    subt = [pricelist.name, f"Currency: {pricelist.currency} · {vat}"] + _validity_lines(pricelist)
    rn = _rate_note(pricelist, pricelist.currency)
    if rn:
        subt.append(rn)
    row = _xl_header(ws, ncols, "PRICELIST", subt)

    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_table_header(ws, row, ncols)
    row += 1

    current = None
    for line in pricelist.lines:
        if line.section and line.section != current:
            current = line.section
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=current)
            c.fill = PatternFill("solid", fgColor="F47A21")
            c.font = Font(bold=True, color="FFFFFF")
            row += 1
        decimals = cx.rounding_for(pricelist.currency)
        fmt = "#,##0" if decimals == 0 else "#,##0." + "0" * decimals
        vals = [line.product.article_no, line.product.barcode or "",
                line.product.description, line.pack_size or line.product.pack_size or ""]
        for col, v in enumerate(vals, start=1):
            cc = ws.cell(row=row, column=col, value=v)
            cc.border = _BORDER
        c = 5
        for t in tiers:
            eff = effective_line_price(line, t.key)
            cc = ws.cell(row=row, column=c)
            if eff["amount"] is not None:
                cc.value = float(eff["amount"])
                cc.number_format = fmt
            if eff["is_fixed"]:
                cc.font = Font(italic=True, color="2B2B2B")
            cc.border = _BORDER
            c += 1
        if has_box:
            for v in (line.box_small, line.box_medium, line.box_large):
                cc = ws.cell(row=row, column=c, value=v)
                cc.border = _BORDER
                c += 1
        row += 1

    _autosize(ws, ncols)
    return _wb_bytes(wb)


def offer_to_excel(offer):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Offer {offer.number}"[:31]
    headers = ["Art No", "Description", "Pack", "Tier", "Qty", "Unit Price", "Disc %", "Line Total"]
    ncols = len(headers)
    vat = _vat_off_label(offer) if not offer.vat_applicable else f"VAT {offer.vat_rate:.0f}%"
    subt = [f"Offer {offer.number} for {offer.customer.name}",
            f"Currency: {offer.currency} · {vat}"] + _validity_lines(offer)
    if offer.currency == "USD" and offer.exchange_rate_value:
        subt.append(f"Rate used (stamped): UGX {float(offer.exchange_rate_value):,.0f} / USD")
    row = _xl_header(ws, ncols, "QUOTATION / QUOTE", subt)

    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_table_header(ws, row, ncols)
    row += 1
    decimals = cx.rounding_for(offer.currency)
    fmt = "#,##0" if decimals == 0 else "#,##0." + "0" * decimals
    for l in offer.lines:
        note = " [fixed]" if l.is_fixed else ""
        ws.cell(row=row, column=1, value=l.article_no or "")
        ws.cell(row=row, column=2, value=(l.description or "") + note)
        ws.cell(row=row, column=3, value=l.pack_size or "")
        ws.cell(row=row, column=4, value=l.tier_label or "")
        ws.cell(row=row, column=5, value=l.quantity)
        c6 = ws.cell(row=row, column=6, value=float(l.unit_price or 0)); c6.number_format = fmt
        ws.cell(row=row, column=7, value=l.discount_pct or 0)
        c8 = ws.cell(row=row, column=8, value=float(l.line_total or 0)); c8.number_format = fmt
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).border = _BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=5, value="Subtotal").font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(offer.subtotal)).number_format = fmt
    row += 1
    if offer.vat_applicable:
        ws.cell(row=row, column=5, value=f"VAT ({offer.vat_rate:.0f}%)").font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(offer.vat_amount)).number_format = fmt
    else:
        # amount cell holds 0 (a number); the label carries the wording
        ws.cell(row=row, column=5, value=f"VAT — {_vat_off_label(offer)}").font = Font(bold=True)
        ws.cell(row=row, column=6, value=0).number_format = fmt
    row += 1
    ws.cell(row=row, column=5, value="TOTAL").font = Font(bold=True, color="ED1C24")
    tc = ws.cell(row=row, column=6, value=float(offer.total)); tc.number_format = fmt
    tc.font = Font(bold=True, color="ED1C24")
    _autosize(ws, ncols)
    return _wb_bytes(wb)


# ---------------------------------------------------------------------------
# small utils
# ---------------------------------------------------------------------------
def _num(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _autosize(ws, ncols):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = 12
        for cell in ws[letter]:
            try:
                if cell.value is not None:
                    width = max(width, min(48, len(str(cell.value)) + 2))
            except Exception:
                pass
        ws.column_dimensions[letter].width = width


def _wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def portal_welcome_pdf(customer, user, temp_pw, portal_url):
    """One-page welcome sheet for a portal login: credentials plus the short
    operations guide. Called with a freshly reset temporary password — the
    stored hash can never be printed, so generating this sheet always goes
    together with a password reset."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            title=f"Portal access - {customer.name}",
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=10 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    elements = []
    _doc_header(elements, styles, "CUSTOMER PORTAL ACCESS", customer.name)

    acc_rows = [("Account", f"<b>{customer.name}</b>"),
                ("Contact", customer.contact_name),
                ("Email", customer.email), ("Phone", customer.phone)]
    login_rows = [("Portal", f"<b>{portal_url}</b>"),
                  ("Username", f"<b>{user.username}</b>"),
                  ("Temp. password", f"<b>{temp_pw}</b>")]
    info = Table([[_info_box("ACCOUNT", acc_rows, styles),
                   _info_box("LOGIN DETAILS", login_rows, styles)]],
                 colWidths=[90 * mm, 91 * mm])
    info.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (0, 0), (0, 0), 0),
                              ("RIGHTPADDING", (1, 0), (1, 0), 0)]))
    elements.append(info)
    elements.append(Spacer(1, 4 * mm))

    note = ParagraphStyle("pw_note", parent=styles["Normal"], fontSize=9,
                          textColor=RF_CHARCOAL, leading=12)
    h = ParagraphStyle("pw_h", parent=styles["Normal"], fontSize=10.5,
                       textColor=RF_INK, fontName="Helvetica-Bold",
                       spaceBefore=8, spaceAfter=2)
    elements.append(Paragraph(
        "Sign in with the details above. The portal asks you to set your own "
        "password immediately; the temporary password stops working after "
        "that. Keep this sheet confidential until then.", note))
    elements.append(Spacer(1, 2 * mm))

    elements.append(Paragraph("How the portal works", h))
    for title, body in [
        ("1. Your pricelist",
         "My Pricelist shows your agreed prices. Prices update automatically "
         "when new lists take effect, so what you see is always current."),
        ("2. Placing an order",
         "Click New Order, pick your products and quantities, and submit. "
         "You receive an order number immediately and our team confirms the "
         "order before it enters fulfilment. You may attach your own LPO."),
        ("3. Tracking orders",
         "The Home page lists your draft, current, and past orders with live "
         "status from confirmation to delivery. Open any order for the "
         "detail or the order PDF."),
        ("4. Messages",
         "Questions or changes on an order go through Messages. Our team "
         "replies in the portal and you see a notification badge."),
        ("5. Offers and promotions",
         "Offers we issue to you appear on your Home page. Accepted offers "
         "convert into orders. Running promotions show there too."),
        ("6. Your account",
         "Change your password any time under Account. A login serving "
         "several outlets picks the outlet in the top bar when ordering."),
    ]:
        elements.append(Paragraph(f"<b>{title}</b> — {body}", note))
        elements.append(Spacer(1, 1.5 * mm))

    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph(
        "Need help? Contact your sales representative or send us a message "
        "in the portal.", note))
    doc.build(elements)
    buf.seek(0)
    return buf.read()
