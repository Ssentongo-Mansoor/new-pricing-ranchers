"""Load historical invoices from the Odoo export 'Invoice List Customer-Wise'.

One sheet, one row per invoice (plus year group-header rows we skip). Columns:
 0 Odoo Invoice Number   2 Customer (Invoice Partner Display)   4 Invoice/Bill Date
 5 Due Date   9 Salesperson   11 Untaxed Amount Signed   12 Total Signed
 13 Currency   14 Payment Status   15 Status   7 Company Type   1 EFRIS

We match each customer name to an existing customer where possible and store one
Invoice row each. Analytics only — no orders created. Clears the pivot
sales_history table since invoices replace it as the history source.

Run: DATABASE_URL=sqlite:////tmp/work.db python import_invoices.py "<xlsx path>"
"""
import re
import sys
import collections
import datetime

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, Invoice


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def run(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    match = build_matcher()
    # clear any previous invoice load (pivot sales_history is kept for product mix)
    db.session.query(Invoice).delete()
    db.session.flush()

    cust_cache = {}
    n = matched = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        num = r[0]
        if num is None:
            continue
        num = str(num)
        if "(" in num and num[:1].isdigit():   # year group header
            continue
        cname = (r[2] or "").strip()
        if cname not in cust_cache:
            cust_cache[cname] = match(cname)
        cid = cust_cache[cname]
        if cid:
            matched += 1
        sp = (r[9] or "").strip()
        sp = re.sub(r"\s*\(.*?\)\s*$", "", sp)   # drop "(sales)" suffixes
        db.session.add(Invoice(
            number=num, customer_id=cid, customer_name=cname, salesperson=sp or None,
            invoice_date=_date(r[4]), due_date=_date(r[5]),
            untaxed=_num(r[11]), total=_num(r[12]),
            currency=(r[13] or "UGX"), payment_status=(r[14] or "").strip() or None,
            company_type=(r[7] or "").strip() or None, efris=str(r[1] or "") or None))
        n += 1
        if n % 20000 == 0:
            db.session.flush()
    db.session.commit()
    n_cust = len(cust_cache)
    print(f"Loaded {n} invoices. Customers: {n_cust} "
          f"({sum(1 for v in cust_cache.values() if v)} matched). Matched invoices: {matched}.")
    # quick validation
    for y in (2024, 2025, 2026):
        rows = db.session.scalars(db.select(Invoice).where(
            db.extract('year', Invoice.invoice_date) == y, Invoice.currency == "UGX")).all()
        tot = sum(float(i.untaxed or 0) for i in rows if i.payment_status != "Reversed")
        print(f"  {y}: {len(rows)} UGX invoices, untaxed (excl reversed) {tot:,.0f}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "imports/Invoice List Customer-Wise 2024, 2025, 2026.xlsx"
    app = create_app()
    with app.app_context():
        run(path)
