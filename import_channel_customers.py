"""Bulk import of the CUSTOMER CHANNEL PRICELIST workbook.

Creates every customer in the sheet and links the generic/customer pricelist(s)
named in the PRICELIST column (allocation). Distributor-channel rows are created
as distributors (separate tab). Category is set from the channel when it maps
cleanly; otherwise left blank for later. Rows whose pricelist name has no match
in the app are still created, with no list linked, and reported.

Usage:
  python import_channel_customers.py            # dry run, prints the plan
  python import_channel_customers.py --commit   # apply to the database
"""
import os
import sys
import collections

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, CustomerCategory, Pricelist
from blueprints.customers import ensure_customer_categories

SHEETFILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "imports", "CUSTOMER CHANNEL PRICELIST.xlsx")

# canonical pricelist names as stored in the DB
N = {
    "horeca":      "Business to Business – Local · HORECA",
    "super":       "Business to Business – Local · Supermarket",
    "dist_horeca": "Business to Distributor – Local · HORECA",
    "dist_super":  "Business to Distributor – Local · Supermarket",
    "exp_horeca":  "Export Business to Business · HORECA",
    "exp_super":   "Export Business to Business · Supermarket",
    "exp_dist_horeca": "Export Business to Distributor · HORECA",
    "exp_dist_super":  "Export Business to Distributor · Supermarket",
    "betar":       "Betar — Local",
    "ranchers_meat": "Ranchers Finest Meat Supermarkets",
    "carrefour":   "Carrefour — Pricelist (April 2026)",
    "capital":     "Capital Shoppers — Pricelist (March 2026)",
    "sokoni":      "Sokoni — Pricelist (April 2026)",
    "gcc":         "GCC — Pricelist (May 2026)",
}


def map_pricelist_keys(text):
    """Return (set_of_keys, matched_any) for a PRICELIST cell string."""
    if not text:
        return set(), False
    s = " ".join(str(text).upper().split())
    keys = set()

    # customer-specific named lists
    if "CARREFOUR" in s: keys.add("carrefour")
    if "CAPITAL" in s:   keys.add("capital")
    if "SOKONI" in s:    keys.add("sokoni")
    if "GCC" in s:       keys.add("gcc")
    if "BETAR" in s:     keys.add("betar")

    # Ranchers own meat supermarkets (consume the phrase so it is not double-read)
    if "RANCHERS" in s and "MEAT" in s:
        keys.add("ranchers_meat")
        s = s.replace("RANCHERS MEAT SUPERMARKETS", " ").replace("RANCHERS MEAT SUPERMARKET", " ")

    export = "EXPORT" in s
    dist = "DISTRIBUTOR" in s
    has_horeca = "HORECA" in s
    has_super = "SUPERMAR" in s   # tolerate SUPERMARET / SUPERMARK / SUPERMARKELIT typos

    if export:
        if dist:
            if has_horeca: keys.add("exp_dist_horeca")
            if has_super:  keys.add("exp_dist_super")
        else:
            if has_horeca: keys.add("exp_horeca")
            if has_super:  keys.add("exp_super")
    else:
        if dist and "LOCAL" in s:
            # business to distributor local — no sub-channel given -> link both
            if has_horeca: keys.add("dist_horeca")
            if has_super:  keys.add("dist_super")
            if not has_horeca and not has_super:
                keys.update({"dist_horeca", "dist_super"})
        else:
            if has_horeca: keys.add("horeca")
            if has_super:  keys.add("super")

    return keys, bool(keys)


CHANNEL_CATEGORY = {
    "SUPERMARKET": "Supermarket", "SUPERMARET": "Supermarket",
    "SUPERMARK": "Supermarket", "SUPERMARKE T": "Supermarket",
    "SCHOOL": "School / Institution",
    "BUTCHERSHOP": "Butchery", "MEAT SHOP": "Butchery", "MEATSHOP": "Butchery",
}


def channel_is_distributor(channel):
    return bool(channel) and "DISTRIBUTOR" in channel.upper()


def channel_is_export(channel):
    return bool(channel) and "EXPORT" in channel.upper()


def channel_category(channel):
    if not channel:
        return None
    up = " ".join(channel.upper().split())
    for frag, cat in CHANNEL_CATEGORY.items():
        if frag in up:
            return cat
    return None


def load_rows():
    wb = load_workbook(SHEETFILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        channel = ws.cell(r, 2).value
        plist = ws.cell(r, 3).value
        if name and str(name).strip():
            out.append((str(name).strip(),
                        str(channel).strip() if channel else None,
                        str(plist).strip() if plist else None))
    return out


def run(commit=False):
    ensure_customer_categories()
    pl_by_name = {p.name: p for p in db.session.scalars(db.select(Pricelist))}
    cat_by_name = {c.name: c for c in db.session.scalars(db.select(CustomerCategory))}

    rows = load_rows()
    created = updated = 0
    n_dist = n_cust = 0
    unmatched = []          # (customer, raw pricelist text)
    link_count = 0
    cat_set = 0

    for name, channel, plist in rows:
        keys, matched = map_pricelist_keys(plist)
        lists = [pl_by_name[N[k]] for k in keys if N[k] in pl_by_name]
        if not matched:
            unmatched.append((name, plist))

        c = db.session.scalar(db.select(Customer).filter_by(name=name))
        if c is None:
            c = Customer(name=name)
            db.session.add(c)
            created += 1
        else:
            updated += 1

        if channel_is_distributor(channel):
            c.segment = "distributor"
            n_dist += 1
        else:
            if not c.segment:
                c.segment = "customer"
            n_cust += 1

        c.market = "export" if channel_is_export(channel) else (c.market or "local")
        if not c.default_currency:
            c.default_currency = "UGX"

        cat = channel_category(channel)
        if cat and cat in cat_by_name:
            c.category_id = cat_by_name[cat].id
            cat_set += 1

        if lists:
            existing = {p.id for p in (c.allowed_pricelists or [])}
            merged = list(c.allowed_pricelists or [])
            for p in lists:
                if p.id not in existing:
                    merged.append(p)
                    existing.add(p.id)
            c.allowed_pricelists = merged
            link_count += len(lists)

        if commit:
            db.session.flush()

    print(f"Rows in sheet: {len(rows)}")
    print(f"Created: {created} | matched existing (updated): {updated}")
    print(f"Distributor rows: {n_dist} | customer rows: {n_cust}")
    print(f"Category set from channel: {cat_set}")
    print(f"Pricelist links made (sum): {link_count}")
    print(f"Rows with NO matching pricelist (created, not linked): {len(unmatched)}")
    print("\nUnmatched pricelist names (distinct counts):")
    for txt, cnt in collections.Counter([u[1] for u in unmatched]).most_common():
        print(f"   {cnt:4}  {txt!r}")

    if commit:
        db.session.commit()
        print("\nCOMMITTED.")
    else:
        db.session.rollback()
        print("\nDRY RUN ONLY (no changes saved). Re-run with --commit to apply.")


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        run(commit="--commit" in sys.argv)
