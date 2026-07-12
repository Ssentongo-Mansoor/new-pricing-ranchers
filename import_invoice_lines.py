"""Load itemized invoice lines from the Odoo export
'Journal Entry (account.move) (customer_invoices_itemized)'.

Layout (one row per invoice LINE):
  0 Invoice Partner Display Name   1 Invoice/Bill Date   2 Number
  3 Invoice lines/Product          4 Invoice lines/Quantity
  5 Invoice lines/Amount in Currency (journal-signed: sales negative)
  6 Total in Currency Signed       7 Total Signed
  8 Untaxed Amount Signed          9 Reference
Rows with a Number start an invoice; continuation rows carry only columns
3-5 and belong to the invoice above. Customer subtotal rows ("NAME (n)",
no number, no product) are skipped.

Behaviour:
  * Headers: existing invoices (matched by number) are kept as they are —
    the account.move export stays the source for salesperson and payment
    status. Missing headers (e.g. the 28 Jun gap day, 2023 history) are
    created from this file with customer matching.
  * Lines: REPLACED per invoice on every run, so re-uploads are idempotent.
  * Amounts: sign flipped once (stored net, sales positive).
  * Product match: Odoo label -> catalogue product via the curated
    sales_history linkage (exact label match).

Run:  SECRET_KEY=x python3 import_invoice_lines.py "<xlsx>"
"""
import re
import sys
import collections
import datetime

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, Invoice, InvoiceLine, SalesHistory


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def build_customer_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def build_product_map():
    """Odoo product label -> catalogue product_id, from the curated
    sales_history linkage (link_history_products.py)."""
    m = {}
    for label, pid in db.session.execute(
            db.select(SalesHistory.product, SalesHistory.product_id)
            .where(SalesHistory.product_id.isnot(None)).distinct()):
        if label and label not in m:
            m[label] = pid
    return m


def run(path):
    cmatch = build_customer_matcher()
    pmap = build_product_map()
    inv_by_number = {i.number: i for i in db.session.scalars(db.select(Invoice))}
    cust_cache = {}

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header

    created_headers = touched = n_lines = matched_lines = 0
    cur = None                     # current Invoice object
    pending = []                   # lines for cur
    replaced = set()               # invoice ids whose lines were cleared this run

    def flush():
        nonlocal n_lines, matched_lines, pending
        if cur is None or not pending:
            pending = []
            return
        if cur.id not in replaced:
            db.session.query(InvoiceLine).filter_by(invoice_id=cur.id).delete(
                synchronize_session=False)
            replaced.add(cur.id)
        for prod, qty, amt in pending:
            pid = pmap.get(prod)
            if pid:
                matched_lines += 1
            db.session.add(InvoiceLine(
                invoice_id=cur.id, product_name=prod, product_id=pid,
                quantity=float(qty) if qty is not None else None,
                amount=-(_num(amt) or 0)))       # journal-signed -> net positive
            n_lines += 1
        pending = []

    for r in rows:
        name, d, num, prod = r[0], _date(r[1]), r[2], r[3]
        if num is not None:
            flush()
            num = str(num).strip()
            inv = inv_by_number.get(num)
            if inv is None:
                cname = (name or "").strip()
                if cname not in cust_cache:
                    cust_cache[cname] = cmatch(cname)
                inv = Invoice(number=num, customer_id=cust_cache[cname],
                              customer_name=cname, invoice_date=d,
                              untaxed=_num(r[8]), total=_num(r[7]),
                              currency="UGX",
                              # NULL fails the dashboards' != 'Reversed'
                              # filter; default until the header export
                              # supplies the real status.
                              payment_status="Not Paid")
                db.session.add(inv)
                db.session.flush()
                inv_by_number[num] = inv
                created_headers += 1
            cur = inv
            touched += 1
            if prod is not None:
                pending.append((str(prod).strip(), r[4], r[5]))
        elif prod is not None and cur is not None:
            pending.append((str(prod).strip(), r[4], r[5]))
        # else: customer subtotal row — ignore
        if n_lines and n_lines % 50000 == 0:
            db.session.flush()
    flush()
    db.session.commit()
    print(f"Invoices touched {touched} (headers created {created_headers}). "
          f"Lines loaded {n_lines}, product-matched {matched_lines} "
          f"({matched_lines * 100 // max(n_lines, 1)}%).")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    app = create_app()
    with app.app_context():
        run(sys.argv[1])
