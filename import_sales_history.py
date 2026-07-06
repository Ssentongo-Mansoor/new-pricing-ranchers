"""Load historical invoiced sales from the yearly pivot export into sales_history.

The export 'Pivot Invoices Analysis ...xlsx' has one sheet per year, each a
two-level pivot: customer (indent 5) -> product (indent 10), with columns
Untaxed Total (UGX, net) and Product Quantity. We flatten it, match each
customer name to an existing customer where possible, and store one row per
customer/product/year. Analytics only — no orders are created.

Run:  DATABASE_URL=sqlite:////tmp/work.db python import_sales_history.py "<xlsx path>"
"""
import re
import sys
import collections

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, SalesHistory

YEARS = ("2024", "2025", "2026")


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _nlead(s):
    return len(s) - len(s.lstrip(" "))


def parse(path):
    wb = load_workbook(path, data_only=True)
    rows = []
    for year in YEARS:
        if year not in wb.sheetnames:
            continue
        ws = wb[year]
        cust = None
        for r in ws.iter_rows(values_only=True):
            a = r[0]
            if a is None:
                continue
            a = str(a)
            lead = _nlead(a)
            name = a.strip()
            if lead == 0 and name == "Total":
                continue
            if lead == 5:
                cust = name
            elif lead == 10 and cust is not None:
                rows.append((int(year), cust, name, float(r[1] or 0), float(r[2] or 0)))
    return rows


def build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def run(path):
    rows = parse(path)
    match = build_matcher()
    db.session.query(SalesHistory).delete()
    db.session.flush()
    matched = 0
    cust_cache = {}
    n = 0
    for year, cust, prod, rev, qty in rows:
        if cust not in cust_cache:
            cust_cache[cust] = match(cust)
        cid = cust_cache[cust]
        if cid:
            matched += 1
        db.session.add(SalesHistory(
            customer_id=cid, customer_name=cust, product=prod, year=year,
            revenue=round(rev, 2), quantity=qty, is_return=(rev < 0)))
        n += 1
    db.session.commit()
    n_cust = len(cust_cache)
    n_matched_cust = sum(1 for v in cust_cache.values() if v)
    print(f"Loaded {n} rows. Customers: {n_cust} ({n_matched_cust} matched, "
          f"{n_cust - n_matched_cust} unmatched). Matched lines: {matched}.")
    for y in YEARS:
        tot = sum(float(s.revenue) for s in db.session.scalars(
            db.select(SalesHistory).filter_by(year=int(y))))
        print(f"  {y}: {tot:,.2f} UGX")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "imports/Pivot Invoices Analysis 2024, 2025, 2026.xlsx"
    app = create_app()
    with app.app_context():
        run(path)
