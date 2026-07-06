"""Excel importer for the Ranchers Finest pricing app.

Two entry points:

* :func:`import_all_seed` — the one-time seed: imports the six source workbooks
  using explicit per-sheet *profiles* (the sheets do not share a layout). The
  Zanzibar and Skylight sheets seed as **customer** pricelists tied to
  placeholder customers.
* :func:`import_mapped_sheet` — the on-demand path used by the admin upload UI:
  the user maps columns to tiers, so new generic lists are added without code.

Both record an :class:`ImportReport` capturing every row that fails to parse.

Run standalone:  ``python importer.py``  (imports the seed workbooks).
"""
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
ART_RE = re.compile(r"^[A-Z]{1,4}\d+[A-Z]*$")

COMMODITY_KEYWORDS = [
    ("BETAR", "Betar"), ("NYAMA", "Betar"), ("BEEFY", "Betar"),
    ("SAUSAGE", "Sausages"), ("GRILLER", "Sausages"), ("HOTDOG", "Hotdogs & Viennas"),
    ("VIENNA", "Hotdogs & Viennas"), ("FRANK", "Hotdogs & Viennas"),
    ("BACON", "Bacon"), ("HAM", "Cold Meats"), ("SALAMI", "Cold Meats"),
    ("COLD", "Cold Meats"), ("PASTRAMI", "Cold Meats"), ("MORTADELLA", "Cold Meats"),
    ("BURGER", "Burgers"), ("PATTY", "Burgers"),
    ("BEEF", "Beef"), ("PORK", "Pork"), ("LAMB", "Lamb"), ("MUTTON", "Lamb"),
    ("GOAT", "Goat"), ("CHICKEN", "Chicken"), ("POULTRY", "Chicken"),
    ("TURKEY", "Turkey"), ("DUCK", "Duck"), ("FISH", "Fish"),
]
QUALITY_WORDS = ["BUDGET", "MARBLED", "PRIME", "CHOICE", "PREMIUM", "BREAKFAST",
                 "BBQ", "COCKTAIL", "STANDARD"]


def is_art(value):
    if value is None:
        return False
    s = str(value).strip().upper()
    return bool(ART_RE.match(s))


def clean_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def clean_art(value):
    s = clean_str(value)
    return s.upper() if s else None


def clean_barcode(value):
    if value is None:
        return None
    if isinstance(value, float):
        return str(int(value))
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def detect_commodity(text):
    if not text:
        return None
    up = text.upper()
    for frag, name in COMMODITY_KEYWORDS:
        if frag in up:
            return name
    return None


def detect_quality(text):
    if not text:
        return None
    up = text.upper()
    for w in QUALITY_WORDS:
        if w in up:
            return w.title()
    return None


def col_letter(idx0):
    return get_column_letter(idx0 + 1)


# ---------------------------------------------------------------------------
# Low-level sheet access
# ---------------------------------------------------------------------------
def _load_rows(path, sheet):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    return rows


def list_sheets(path):
    wb = load_workbook(path, read_only=True)
    return wb.sheetnames


def preview_sheet(path, sheet, max_rows=15):
    rows = _load_rows(path, sheet)
    ncols = max((len(r) for r in rows[:60]), default=0)
    sample = []
    for i, r in enumerate(rows[:max_rows], start=1):
        padded = [r[c] if c < len(r) else None for c in range(ncols)]
        sample.append({"row": i, "cells": padded})
    return {"ncols": ncols, "rows": sample}


def _cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


# ---------------------------------------------------------------------------
# Catalogue upserts (imported lazily to avoid app/db import cycles)
# ---------------------------------------------------------------------------
def _get_db():
    from extensions import db
    return db


def get_or_create_category(name, parent_name=None):
    from models import Category
    db = _get_db()
    parent = None
    if parent_name:
        parent = db.session.scalar(
            db.select(Category).filter_by(name=parent_name, parent_id=None))
        if parent is None:
            parent = Category(name=parent_name)
            db.session.add(parent)
            db.session.flush()
    cat = db.session.scalar(
        db.select(Category).filter_by(name=name,
                                      parent_id=(parent.id if parent else None)))
    if cat is None:
        cat = Category(name=name, parent_id=(parent.id if parent else None))
        db.session.add(cat)
        db.session.flush()
    return cat


def get_or_create_product(article_no, description, barcode=None, pack_size=None,
                          uom=None, category=None):
    from models import Product
    db = _get_db()
    p = db.session.scalar(db.select(Product).filter_by(article_no=article_no))
    if p is None:
        p = Product(article_no=article_no, description=description or article_no,
                    barcode=barcode, pack_size=pack_size, unit_of_measure=uom,
                    category_id=(category.id if category else None))
        db.session.add(p)
        db.session.flush()
    else:
        if not p.barcode and barcode:
            p.barcode = barcode
        if not p.pack_size and pack_size:
            p.pack_size = pack_size
        if not p.category_id and category:
            p.category_id = category.id
    return p


def find_product_by_desc(description):
    from models import Product
    db = _get_db()
    norm = re.sub(r"\s+", " ", (description or "").strip().lower())
    for p in db.session.scalars(_get_db().select(Product)):
        if re.sub(r"\s+", " ", p.description.strip().lower()) == norm:
            return p
    return None


def get_or_create_customer(name, market="export", currency="USD"):
    from models import Customer
    db = _get_db()
    c = db.session.scalar(db.select(Customer).filter_by(name=name))
    if c is None:
        c = Customer(name=name, market=market, default_currency=currency,
                     notes="Placeholder customer created during seed import.")
        db.session.add(c)
        db.session.flush()
    return c


# ---------------------------------------------------------------------------
# Core: import one sheet given a profile/mapping
# ---------------------------------------------------------------------------
def _category_for(section_text, description, commodity_state, quality_state):
    """Resolve a (category) object for a product from section context."""
    commodity = detect_commodity(section_text) or commodity_state \
        or detect_commodity(description)
    quality = detect_quality(section_text) or quality_state
    if not commodity:
        return None, commodity_state, quality_state
    if quality and quality.upper() not in ("STANDARD",):
        cat = get_or_create_category(quality, parent_name=commodity)
    else:
        cat = get_or_create_category(commodity)
    return cat, commodity, quality


def _import_sheet_core(rows, mapping, pricelist, report, synth_prefix=None):
    """Shared row loop. ``mapping`` keys: data_start (1-based), art_col, barcode_col,
    desc_col, pack_col, units_col, box_*_col, tiers [(key,label,col)], art_required.
    """
    from models import PricelistTier, PricelistLine, LinePrice
    db = _get_db()

    # tiers
    tier_objs = {}
    for order, (key, label, _col) in enumerate(mapping["tiers"]):
        t = PricelistTier(pricelist_id=pricelist.id, key=key, label=label, sort_order=order)
        db.session.add(t)
        db.session.flush()
        tier_objs[key] = t

    data_start = mapping["data_start"]
    art_required = mapping.get("art_required", True)
    desc_col = mapping["desc_col"]
    art_col = mapping.get("art_col")

    current_section = None
    commodity_state = None
    quality_state = None
    sort_order = 0
    synth_n = 0

    # Scan from the top so section/category banners that sit *above* the first
    # data row (e.g. "BEEF", "CATEGORY 1 - BUDGET BEEF") are captured, but only
    # accept actual data rows at or after data_start.
    for i in range(0, len(rows)):
        row = rows[i]
        rownum = i + 1
        art_val = _cell(row, art_col) if art_col is not None else None
        desc_val = clean_str(_cell(row, desc_col))

        # Is this a data row?
        if rownum < data_start:
            data_row = False
        elif art_required:
            data_row = is_art(art_val)
        else:
            data_row = desc_val is not None and to_number(_cell(row, mapping["tiers"][0][2])) is not None

        if not data_row:
            # treat as a section header if it carries text
            txt = desc_val or clean_str(art_val)
            if txt and not _looks_like_header_noise(txt):
                current_section = txt
                c = detect_commodity(txt)
                if c:
                    if c != commodity_state:
                        quality_state = None   # don't carry quality across commodities
                    commodity_state = c
                q = detect_quality(txt)
                if q:
                    quality_state = q
            continue

        # ---- build/find product ----
        if art_required:
            article_no = clean_art(art_val)
            description = desc_val or article_no
        else:
            description = desc_val
            existing = find_product_by_desc(description)
            if existing:
                article_no = existing.article_no
            else:
                synth_n += 1
                article_no = f"{synth_prefix}{synth_n:03d}"

        if not description:
            report["failed"] += 1
            report["lines"].append(f"Row {rownum}: missing description (art {art_val}).")
            continue

        pack = clean_str(_cell(row, mapping.get("pack_col"))) if mapping.get("pack_col") is not None else None
        units = clean_str(_cell(row, mapping.get("units_col"))) if mapping.get("units_col") is not None else None
        barcode = clean_barcode(_cell(row, mapping.get("barcode_col"))) if mapping.get("barcode_col") is not None else None

        cat, commodity_state, quality_state = _category_for(
            current_section, description, commodity_state, quality_state)

        product = get_or_create_product(
            article_no, description, barcode=barcode,
            pack_size=pack or units, uom=_guess_uom(pack or units, pricelist.channel),
            category=cat)

        # ---- collect tier prices ----
        prices = {}
        any_price = False
        for key, label, col in mapping["tiers"]:
            val = to_number(_cell(row, col))
            prices[key] = val
            if val is not None:
                any_price = True
        if not any_price:
            report["failed"] += 1
            report["lines"].append(
                f"Row {rownum}: {article_no} '{description}' had no numeric price; skipped.")
            continue

        # auto-fill Incl VAT from Excl when missing on a VAT list
        if pricelist.vat_applicable:
            if "incl_vat" in prices and prices.get("incl_vat") is None \
               and prices.get("excl_vat") is not None:
                prices["incl_vat"] = round(prices["excl_vat"] * (1 + pricelist.vat_rate / 100.0), 2)
            if "excl_vat" in prices and prices.get("excl_vat") is None \
               and prices.get("incl_vat") is not None:
                prices["excl_vat"] = round(prices["incl_vat"] / (1 + pricelist.vat_rate / 100.0), 2)

        sort_order += 1
        line = PricelistLine(
            pricelist_id=pricelist.id, product_id=product.id,
            section=current_section, pack_size=pack or units,
            units_per_pack=units,
            box_small=to_number(_cell(row, mapping.get("box_small_col"))) if mapping.get("box_small_col") is not None else None,
            box_medium=to_number(_cell(row, mapping.get("box_medium_col"))) if mapping.get("box_medium_col") is not None else None,
            box_large=to_number(_cell(row, mapping.get("box_large_col"))) if mapping.get("box_large_col") is not None else None,
            sort_order=sort_order)
        db.session.add(line)
        db.session.flush()
        for key, _label, _col in mapping["tiers"]:
            db.session.add(LinePrice(line_id=line.id, tier_id=tier_objs[key].id,
                                     amount=prices.get(key)))
        report["ok"] += 1

    return pricelist


def _looks_like_header_noise(txt):
    up = txt.upper()
    noise = ["ALL PRICES", "PRICELIST", "PACKING PER BOX", "PRODUCTS",
             "VALID FOR", "ZERO RATED", "ZERO-RATED", "PRODUCT DESCRIPTION",
             "ART NO", "BARCODE", "UNITS / PACK", "PACK SIZE", "BOX SIZE",
             "HORECA CHANNEL", "CODE", "KEY ASSUMPTIONS", "EXCHANGE RATE",
             "AIRFREIGHT", "AWB", "CIF DEFINITION", "IMPORT DUTY",
             "DISTRIBUTOR MARGIN", "INSURANCE", "CORRECTED VERSION"]
    return any(n in up for n in noise)


def _guess_uom(pack, channel):
    if pack:
        up = pack.upper()
        if up.strip() in ("KG", "PCS", "PC"):
            return up.strip().lower()
    if channel == "horeca":
        return "kg"
    if channel == "supermarket":
        return "pack"
    return None


# ---------------------------------------------------------------------------
# Generic (mapped) import for the upload UI
# ---------------------------------------------------------------------------
def import_mapped_sheet(path, sheet, mapping, meta, source_label=None):
    from models import Pricelist, ImportReport
    from services import settings as settings_svc
    db = _get_db()
    rows = _load_rows(path, sheet)

    tiers = []
    for col, label in zip(mapping["tier_cols"], mapping["tier_labels"]):
        key = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_") or f"tier_{col}"
        tiers.append((key, label, col))

    customer_id = meta.get("customer_id")
    pl = Pricelist(
        name=meta["name"], channel=meta.get("channel", "mixed"),
        market=meta.get("market", "local"), currency=meta.get("currency", "UGX"),
        vat_applicable=meta.get("vat_applicable", True),
        vat_rate=settings_svc.get_float("vat_rate", 18.0),
        is_customer=bool(meta.get("is_customer")),
        customer_id=int(customer_id) if customer_id else None,
        source_file=source_label or os.path.basename(path))
    db.session.add(pl)
    db.session.flush()

    core_map = {
        "data_start": mapping["data_start"],
        "art_col": mapping.get("art_col"),
        "barcode_col": mapping.get("barcode_col"),
        "desc_col": mapping["desc_col"],
        "pack_col": mapping.get("pack_col"),
        "units_col": mapping.get("pack_col"),
        "box_small_col": mapping.get("box_small_col"),
        "box_medium_col": mapping.get("box_medium_col"),
        "box_large_col": mapping.get("box_large_col"),
        "tiers": tiers,
        "art_required": mapping.get("art_col") is not None,
    }
    report = {"ok": 0, "failed": 0, "lines": []}
    _import_sheet_core(rows, core_map, pl, report,
                       synth_prefix=_synth_prefix(meta["name"]))

    rep = ImportReport(source=f"{source_label or os.path.basename(path)} :: {sheet}",
                       rows_ok=report["ok"], rows_failed=report["failed"],
                       detail="\n".join(report["lines"]))
    db.session.add(rep)
    return pl, report


def _synth_prefix(name):
    letters = re.sub(r"[^A-Za-z]", "", name).upper()
    return (letters[:3] or "GEN")


# ---------------------------------------------------------------------------
# Seed profiles for the six source workbooks
# ---------------------------------------------------------------------------
def _find_file(seed_dir, *needles):
    for f in os.listdir(seed_dir):
        up = f.upper()
        if all(n.upper() in up for n in needles):
            return os.path.join(seed_dir, f)
    return None


def seed_profiles(seed_dir):
    """Return the ordered list of (file_path, sheet, mapping, meta) to import."""
    profiles = []

    b2b = _find_file(seed_dir, "BUSINESS TO BUSINESS LOCAL")
    dist = _find_file(seed_dir, "BUSINESS TO DISTRIBUTOR LOCAL")
    exb2b = _find_file(seed_dir, "EXPORT BUSINESS TO BUSINESS")
    exdist = _find_file(seed_dir, "EXPORT BUSINESS TO DISTRIBUTOR")
    meat = _find_file(seed_dir, "MEAT SUPERMARKETS")
    betar = _find_file(seed_dir, "BETAR")

    def horeca_local(art=0, desc=1, units=2, bs=4, bm=6, bl=8, t1=10, t2=11,
                     l1="Price Excl VAT", l2="Price Incl VAT",
                     k1="excl_vat", k2="incl_vat", data_start=12):
        return {
            "data_start": data_start, "art_col": art, "desc_col": desc,
            "units_col": units, "pack_col": units,
            "box_small_col": bs, "box_medium_col": bm, "box_large_col": bl,
            "tiers": [(k1, l1, t1), (k2, l2, t2)], "art_required": True,
        }

    def supermarket(art=0, barcode=1, desc=2, pack=3, tiers=None, data_start=6,
                    bs=7, bm=8, bl=9):
        return {
            "data_start": data_start, "art_col": art, "barcode_col": barcode,
            "desc_col": desc, "pack_col": pack,
            "box_small_col": bs, "box_medium_col": bm, "box_large_col": bl,
            "tiers": tiers, "art_required": True,
        }

    def export_horeca(data_start=7):
        return {
            "data_start": data_start, "art_col": 0, "desc_col": 1,
            "units_col": 2, "pack_col": 2,
            "tiers": [("price_kg", "Price per kg", 3)], "art_required": True,
        }

    def export_super(data_start=6):
        return {
            "data_start": data_start, "art_col": 0, "barcode_col": 1,
            "desc_col": 2, "pack_col": 3,
            "tiers": [("price_pack", "Price per pack", 4)], "art_required": True,
        }

    # 1. BUSINESS TO BUSINESS LOCAL
    if b2b:
        profiles.append((b2b, "HORECA ", horeca_local(),
                         dict(name="Business to Business – Local · HORECA",
                              channel="horeca", market="local", currency="UGX",
                              vat_applicable=True)))
        profiles.append((b2b, "SUPERMARKET",
                         supermarket(tiers=[("excl_vat", "Price Excl VAT", 4),
                                            ("incl_vat", "Price Incl VAT", 5),
                                            ("rrp", "RRP", 6)], data_start=6),
                         dict(name="Business to Business – Local · Supermarket",
                              channel="supermarket", market="local", currency="UGX",
                              vat_applicable=True)))

    # 2. BUSINESS TO DISTRIBUTOR LOCAL
    if dist:
        profiles.append((dist, "HORECA DISTRIBUTOR LOCAL",
                         horeca_local(t1=10, t2=11, l1="Distributor Price",
                                      l2="Distributor Sells At",
                                      k1="dist_price", k2="dist_sells"),
                         dict(name="Business to Distributor – Local · HORECA",
                              channel="horeca", market="local", currency="UGX",
                              vat_applicable=True)))
        profiles.append((dist, "SUPERMARKET DISTRIBUTOR LOCAL",
                         supermarket(tiers=[("dist_price", "Distributor Price", 4),
                                            ("wholesale", "Wholesale Distributor", 5),
                                            ("retail", "Retail", 6)], data_start=5),
                         dict(name="Business to Distributor – Local · Supermarket",
                              channel="supermarket", market="local", currency="UGX",
                              vat_applicable=True)))

    # 3. EXPORT BUSINESS TO BUSINESS
    if exb2b:
        profiles.append((exb2b, "HORECA EXPORT", export_horeca(),
                         dict(name="Export Business to Business · HORECA",
                              channel="horeca", market="export", currency="USD",
                              vat_applicable=False)))
        profiles.append((exb2b, "SUPERMARKET  EXPORT", export_super(),
                         dict(name="Export Business to Business · Supermarket",
                              channel="supermarket", market="export", currency="USD",
                              vat_applicable=False)))

    # 4. EXPORT BUSINESS TO DISTRIBUTOR (incl. customer-specific Zanzibar & Skylight)
    if exdist:
        profiles.append((exdist, "EXPORT DISTRIBUTOR HORECA", export_horeca(),
                         dict(name="Export Business to Distributor · HORECA",
                              channel="horeca", market="export", currency="USD",
                              vat_applicable=False)))
        profiles.append((exdist, "EXPORT DISTRIBUTOR SUPERMARKET", export_super(),
                         dict(name="Export Business to Distributor · Supermarket",
                              channel="supermarket", market="export", currency="USD",
                              vat_applicable=False)))
        profiles.append((exdist, "ZANZIBAR EXPORT DISTRIBUTOR",
                         {"data_start": 14, "art_col": 0, "desc_col": 1,
                          "tiers": [("ex_works", "Ex-Works USD/kg", 2),
                                    ("cif_usd", "CIF USD/kg", 4),
                                    ("wholesale_usd", "Wholesale USD/kg +20%", 6)],
                          "art_required": True},
                         dict(name="Zanzibar Distributor — Export HORECA",
                              channel="horeca", market="export", currency="USD",
                              vat_applicable=False, is_customer=True,
                              customer_name="Zanzibar Distributor",
                              notes=("Customer-specific export list. Key assumptions: "
                                     "TZS/USD 2,605; airfreight USD 1.10/kg; 0% EAC duty; "
                                     "distributor margin 20%."))))
        profiles.append((exdist, "SKYLIGHT PRICELIST",
                         {"data_start": 6, "art_col": None, "desc_col": 0,
                          "tiers": [("price_kg", "Price per kg", 1)],
                          "art_required": False},
                         dict(name="Skylight — Export Pricelist",
                              channel="mixed", market="export", currency="USD",
                              vat_applicable=False, is_customer=True,
                              customer_name="Skylight",
                              notes="Customer-specific export list. All prices USD/kg.")))

    # 5. MEAT SUPERMARKETS
    if meat:
        profiles.append((meat, "LUGOGO AND KABALAGALA",
                         supermarket(tiers=[("excl_vat", "Price Excl VAT", 4),
                                            ("incl_vat", "Price Incl VAT", 5)],
                                     data_start=6, bs=None, bm=None, bl=None),
                         dict(name="Meat Supermarkets · Lugogo & Kabalagala",
                              channel="supermarket", market="local", currency="UGX",
                              vat_applicable=True)))

    # 6. BETAR
    if betar:
        profiles.append((betar, "Local ",
                         {"data_start": 5, "art_col": 0, "barcode_col": 1,
                          "desc_col": 2, "units_col": 3, "pack_col": 4,
                          "tiers": [("excl_vat", "Price Excl VAT", 5),
                                    ("incl_vat", "Price Incl VAT", 6),
                                    ("rrp", "RRP", 7)], "art_required": True},
                         dict(name="Betar — Local", channel="supermarket",
                              market="local", currency="UGX", vat_applicable=True)))
        profiles.append((betar, "Distributor",
                         {"data_start": 5, "art_col": 0, "barcode_col": 1,
                          "desc_col": 2, "units_col": 3, "pack_col": 4,
                          "tiers": [("excl_vat", "Price Excl VAT", 5),
                                    ("incl_vat", "Price Incl VAT", 6)],
                          "art_required": True},
                         dict(name="Betar — Distributor", channel="supermarket",
                              market="local", currency="UGX", vat_applicable=True)))

    return profiles


def import_all_seed(seed_dir, verbose=True):
    """Import all six workbooks. Idempotent: skips a list already imported by name."""
    from datetime import date, timedelta
    from models import Pricelist, ImportReport
    from services import settings as settings_svc
    db = _get_db()

    profiles = seed_profiles(seed_dir)
    if not profiles:
        raise FileNotFoundError(f"No source workbooks found in {seed_dir}")

    totals = {"ok": 0, "failed": 0, "lists": 0}
    for path, sheet, mapping, meta in profiles:
        existing = db.session.scalar(db.select(Pricelist).filter_by(name=meta["name"]))
        if existing:
            if verbose:
                print(f"  = skip (exists): {meta['name']}")
            continue

        customer = None
        if meta.get("is_customer") and meta.get("customer_name"):
            customer = get_or_create_customer(meta["customer_name"],
                                              market=meta.get("market", "export"),
                                              currency=meta.get("currency", "USD"))
        valid_until = None
        if meta.get("market") == "export":
            valid_until = date.today() + timedelta(days=30)

        pl = Pricelist(
            name=meta["name"], channel=meta.get("channel"), market=meta.get("market"),
            currency=meta.get("currency", "UGX"),
            vat_applicable=meta.get("vat_applicable", True),
            vat_rate=settings_svc.get_float("vat_rate", 18.0),
            effective_date=date.today(), valid_until=valid_until,
            notes=meta.get("notes"),
            is_customer=bool(meta.get("is_customer")),
            customer_id=(customer.id if customer else None),
            source_file=f"{os.path.basename(path)} :: {sheet}")
        db.session.add(pl)
        db.session.flush()

        report = {"ok": 0, "failed": 0, "lines": []}
        rows = _load_rows(path, sheet)
        mapping = dict(mapping)
        mapping.setdefault("art_required", True)
        _import_sheet_core(rows, mapping, pl, report,
                           synth_prefix=("SKY" if "Skylight" in meta["name"] else "GEN"))

        rep = ImportReport(source=pl.source_file, rows_ok=report["ok"],
                           rows_failed=report["failed"],
                           detail="\n".join(report["lines"]))
        db.session.add(rep)
        db.session.commit()
        totals["ok"] += report["ok"]
        totals["failed"] += report["failed"]
        totals["lists"] += 1
        if verbose:
            tag = "[customer] " if pl.is_customer else ""
            print(f"  + {tag}{pl.name}: {report['ok']} rows ok, {report['failed']} skipped")

    # Also build a UGX-denominated Skylight list (converted at the rate in force).
    build_skylight_ugx(verbose=verbose)

    if verbose:
        print(f"\nSeed import complete: {totals['lists']} lists, "
              f"{totals['ok']} rows ok, {totals['failed']} skipped.")
    return totals


def build_skylight_ugx(verbose=True, rate_value=None):
    """Create a UGX copy of the Skylight customer pricelist, converting each
    USD/kg price to UGX at the exchange rate in force. Idempotent."""
    from datetime import date, timedelta
    from models import Pricelist, PricelistTier, PricelistLine, LinePrice
    from services import currency as cx
    db = _get_db()

    src = db.session.scalar(
        db.select(Pricelist).filter_by(name="Skylight — Export Pricelist"))
    if src is None:
        if verbose:
            print("  = Skylight (USD) list not found; skip UGX build.")
        return None

    target_name = "Skylight — Pricelist (UGX)"
    if db.session.scalar(db.select(Pricelist).filter_by(name=target_name)):
        if verbose:
            print(f"  = skip (exists): {target_name}")
        return None

    if rate_value is None:
        rate = cx.get_rate("USD")
        if rate is None:
            if verbose:
                print("  ! No UGX→USD rate on file; cannot build Skylight UGX. "
                      "Set a rate, then re-run.")
            return None
        rate_value = rate.rate

    ugx = Pricelist(
        name=target_name, channel=src.channel, market=src.market,
        currency="UGX", vat_applicable=src.vat_applicable,
        vat_rate=src.vat_rate, effective_date=date.today(),
        valid_until=date.today() + timedelta(days=30),
        notes=(f"UGX version of the Skylight list, converted from USD at "
               f"UGX {float(rate_value):,.0f}/USD on {date.today()}."),
        is_customer=True, customer_id=src.customer_id,
        base_pricelist_id=src.id, source_file="derived from Skylight (USD)")
    db.session.add(ugx)
    db.session.flush()
    tier = PricelistTier(pricelist_id=ugx.id, key="price_kg",
                         label="Price per kg", sort_order=0)
    db.session.add(tier)
    db.session.flush()

    n = 0
    for line in src.lines:
        usd = line.price_for("price_kg")
        nl = PricelistLine(pricelist_id=ugx.id, product_id=line.product_id,
                           section=line.section, pack_size=line.pack_size,
                           units_per_pack=line.units_per_pack,
                           sort_order=line.sort_order)
        db.session.add(nl)
        db.session.flush()
        converted = None
        if usd is not None:
            converted = cx.convert(usd, "USD", "UGX", rate_value=rate_value)
        db.session.add(LinePrice(line_id=nl.id, tier_id=tier.id, amount=converted))
        n += 1
    db.session.commit()
    if verbose:
        print(f"  + [customer] {target_name}: {n} lines at UGX {float(rate_value):,.0f}/USD")
    return ugx


def apply_skylight_revision(path, verbose=True):
    """Apply a revised-prices workbook (UGX/kg) to the Skylight UGX list.

    Layout: header row 4, data from row 5 — col A art no, col B description,
    col D revised UGX price. Matched article numbers are updated (logged as a
    price change); unmatched ones are added as new lines.
    """
    from datetime import date, datetime
    from models import (Pricelist, PricelistTier, PricelistLine, LinePrice)
    from services.audit import log
    db = _get_db()

    ugx = db.session.scalar(
        db.select(Pricelist).filter_by(name="Skylight — Pricelist (UGX)"))
    if ugx is None:
        ugx = build_skylight_ugx(verbose=verbose)
    if ugx is None:
        # No USD base to derive from — create a bare UGX Skylight list.
        cust = get_or_create_customer("Skylight", market="export", currency="UGX")
        ugx = Pricelist(name="Skylight — Pricelist (UGX)", channel="mixed",
                        market="export", currency="UGX", vat_applicable=False,
                        is_customer=True, customer_id=cust.id,
                        source_file="Skylight revision")
        db.session.add(ugx)
        db.session.flush()

    tier = db.session.scalar(
        db.select(PricelistTier).filter_by(pricelist_id=ugx.id, key="price_kg"))
    if tier is None:
        tier = PricelistTier(pricelist_id=ugx.id, key="price_kg",
                             label="Price per kg", sort_order=0)
        db.session.add(tier)
        db.session.flush()

    rows = _load_rows(path, _load_first_sheet(path))
    data_start, art_col, desc_col, price_col = 5, 0, 1, 3
    section = commodity = quality = None
    by_pid = {l.product_id: l for l in ugx.lines}
    updated = added = skipped = 0
    report_lines = []

    for i in range(0, len(rows)):
        row = rows[i]
        rownum = i + 1
        art = _cell(row, art_col)
        desc = clean_str(_cell(row, desc_col))
        if rownum < data_start or not is_art(art):
            txt = desc or clean_str(art)
            if txt and not _looks_like_header_noise(txt):
                section = txt
                c = detect_commodity(txt)
                if c:
                    if c != commodity:
                        quality = None
                    commodity = c
                q = detect_quality(txt)
                if q:
                    quality = q
            continue

        price = to_number(_cell(row, price_col))
        if price is None:
            skipped += 1
            report_lines.append(f"Row {rownum}: {art} had no revised price; skipped.")
            continue

        article = clean_art(art)
        cat, commodity, quality = _category_for(section, desc, commodity, quality)
        product = get_or_create_product(article, desc, category=cat)
        line = by_pid.get(product.id)
        if line:
            lp = db.session.scalar(
                db.select(LinePrice).filter_by(line_id=line.id, tier_id=tier.id))
            old = lp.amount if lp else None
            if lp is None:
                db.session.add(LinePrice(line_id=line.id, tier_id=tier.id, amount=price))
            else:
                lp.amount = price
            log("price_change", "pricelist_line", line.id, field=tier.label,
                old_value=old, new_value=price,
                detail=f"{article} — Skylight UGX revision")
            updated += 1
        else:
            mx = max([l.sort_order for l in ugx.lines], default=0)
            nl = PricelistLine(pricelist_id=ugx.id, product_id=product.id,
                               section=section or "REVISED JUNE 2026",
                               pack_size=product.pack_size, sort_order=mx + 1)
            db.session.add(nl)
            db.session.flush()
            db.session.add(LinePrice(line_id=nl.id, tier_id=tier.id, amount=price))
            by_pid[product.id] = nl
            ugx.lines.append(nl)
            log("customer_pricelist_edit", "pricelist", ugx.id,
                detail=f"added {article} (Skylight UGX revision)")
            added += 1

    ugx.effective_date = date.today()
    ugx.notes = (ugx.notes or "") + \
        f"\nRevised prices applied {date.today()} from {os.path.basename(path)}."
    db.session.commit()
    if verbose:
        print(f"Skylight UGX revision: {updated} updated, {added} added, {skipped} skipped.")
    return {"updated": updated, "added": added, "skipped": skipped, "lines": report_lines}


def _load_first_sheet(path):
    return list_sheets(path)[0]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _run_cli():
    from app import create_app
    from config import Config
    app = create_app()
    with app.app_context():
        seed_dir = Config.SEED_DIR
        if len(sys.argv) > 1:
            seed_dir = sys.argv[1]
        print(f"Importing seed workbooks from: {seed_dir}")
        import_all_seed(seed_dir)


if __name__ == "__main__":
    _run_cli()
