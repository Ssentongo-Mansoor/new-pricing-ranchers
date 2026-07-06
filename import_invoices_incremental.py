"""Incremental invoice/credit-note top-up from the Odoo 'Journal Entry
(account.move)' export. UPSERTS by invoice number — never clears the table —
so periodic uploads keep the invoice history continuous.

Handles both layouts:
  * Invoices file:  0 Number  1 FDN  2 Partner  4 Invoice/Bill Date  5 Due
      7 Company Type  9 Salesperson  11 Untaxed Signed  12 Total Signed
      13 Currency  14 Payment Status  18 EFRIS Status
  * Credit-notes file:  0 Number  1 Partner  3 Invoice/Bill Date  4 Due
      6 Untaxed Signed  7 Total Signed  9 Payment Status  11 EFRIS Status
The layout is detected from the header row. Date-group header rows
("29 Jun 2026 (95)") are skipped. Existing numbers are UPDATED (payment
status moves Paid/Not Paid over time); new numbers are inserted.

Run:  SECRET_KEY=x python3 import_invoices_incremental.py "<xlsx>" [more.xlsx ...]
"""
import re
import sys
import collections
import datetime

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, Invoice


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def detect_layout(header):
    """Map field -> column index from the header row."""
    idx = {str(h or "").strip(): i for i, h in enumerate(header)}
    def col(*names):
        for nm in names:
            if nm in idx:
                return idx[nm]
        return None
    return {
        "number": col("Number"),
        "partner": col("Invoice Partner Display Name"),
        "date": col("Invoice/Bill Date"),
        "due": col("Due Date"),
        "untaxed": col("Untaxed Amount Signed"),
        "total": col("Total Signed"),
        "currency": col("Currency"),
        "paystat": col("Payment Status"),
        "company": col("Company Type"),
        "salesperson": col("Salesperson"),
        "efris": col("FDN", "EFRIS Status", "EFRIS"),
    }


def run(paths):
    match = build_matcher()
    existing = {i.number: i for i in db.session.scalars(db.select(Invoice))}
    cust_cache = {}
    ins = upd = skipped = 0
    dates = []

    for path in paths:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Sheet1"]
        rows = ws.iter_rows(values_only=True)
        lay = detect_layout(next(rows))
        if lay["number"] is None or lay["date"] is None:
            print(f"SKIPPING {path}: header not recognised.")
            continue
        for r in rows:
            num = r[lay["number"]]
            if num is None:
                continue
            num = str(num).strip()
            if "(" in num and num[:1].isdigit():   # date group header
                continue
            d = _date(r[lay["date"]])
            if d is None:
                skipped += 1
                continue
            dates.append(d)
            cname = (r[lay["partner"]] or "").strip()
            if cname not in cust_cache:
                cust_cache[cname] = match(cname)
            sp = None
            if lay["salesperson"] is not None:
                sp = re.sub(r"\s*\(.*?\)\s*$", "", (r[lay["salesperson"]] or "").strip()) or None
            vals = dict(
                customer_id=cust_cache[cname], customer_name=cname,
                salesperson=sp, invoice_date=d, due_date=_date(r[lay["due"]]),
                untaxed=_num(r[lay["untaxed"]]), total=_num(r[lay["total"]]),
                currency=(r[lay["currency"]] if lay["currency"] is not None else None) or "UGX",
                payment_status=((r[lay["paystat"]] or "").strip() or None)
                    if lay["paystat"] is not None else None,
                company_type=((r[lay["company"]] or "").strip() or None)
                    if lay["company"] is not None else None,
                efris=(str(r[lay["efris"]] or "") or None)
                    if lay["efris"] is not None else None,
            )
            inv = existing.get(num)
            if inv is not None:
                for k, v in vals.items():
                    if v is not None or k in ("payment_status",):
                        setattr(inv, k, v)
                upd += 1
            else:
                inv = Invoice(number=num, **vals)
                db.session.add(inv)
                existing[num] = inv
                ins += 1
    db.session.commit()
    lo, hi = (min(dates), max(dates)) if dates else (None, None)
    n_match = sum(1 for v in cust_cache.values() if v)
    print(f"Inserted {ins}, updated {upd}, skipped {skipped} (no date). "
          f"Dates {lo} .. {hi}. Customers seen {len(cust_cache)}, matched {n_match}.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    app = create_app()
    with app.app_context():
        run(sys.argv[1:])
