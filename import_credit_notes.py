"""Load credit notes from 'Credit Notes Customer List' into the invoice table
(negative amounts) so revenue per customer is net of credits.

Columns: 0 Number(RINV)  1 Customer  3 Invoice/Bill Date  4 Due Date
 6 Untaxed Amount Signed (neg)  7 Total Signed (neg)  9 Payment Status  10 Status  11 EFRIS

Idempotent: clears existing RINV rows first. Run AFTER import_invoices.py
(which clears the whole invoice table).

Run: DATABASE_URL=sqlite:////tmp/work.db python import_credit_notes.py "<xlsx path>"
"""
import re
import sys
import collections
import datetime

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, Invoice


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def run(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    match = build_matcher()
    db.session.query(Invoice).filter(Invoice.number.like("RINV%")).delete(synchronize_session=False)
    db.session.flush()
    cache = {}
    n = matched = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        num = r[0]
        if num is None:
            continue
        num = str(num)
        if "(" in num and num[:1].isdigit():
            continue
        cname = (r[1] or "").strip()
        if cname not in cache:
            cache[cname] = match(cname)
        cid = cache[cname]
        if cid:
            matched += 1
        db.session.add(Invoice(
            number=num, customer_id=cid, customer_name=cname, salesperson=None,
            invoice_date=_date(r[3]), due_date=_date(r[4]),
            untaxed=_num(r[6]), total=_num(r[7]), currency="UGX",
            payment_status=(r[9] or "").strip() or None,
            company_type="Credit note", efris=str(r[11] or "") or None))
        n += 1
    db.session.commit()
    print(f"Loaded {n} credit notes ({matched} matched to customers).")
    for y in (2024, 2025, 2026):
        tot = sum(float(i.untaxed or 0) for i in db.session.scalars(
            db.select(Invoice).where(Invoice.number.like("RINV%"),
                                     db.extract('year', Invoice.invoice_date) == y)))
        print(f"  {y}: credit notes untaxed {tot:,.0f}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "imports/Credit Notes Customer List (2024, 2025, 2026).xlsx"
    app = create_app()
    with app.app_context():
        run(path)
