"""Import opening stock from 'RF Stock Update.xlsx':
- Coldroom Stock  -> set on-hand of matching catalogue products (sellable stock).
- Dry Store / Blast Freezer / Mini Blast Store / Carcus Chiller -> StoreItems.

Run:  python import_stock.py
"""
import os
from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Product, Store, StoreItem
from services import stock as stock_svc

FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    "imports", "RF Stock Update.xlsx")


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _art(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    return s or None


def import_coldroom(ws):
    by_art = {p.article_no.upper(): p for p in db.session.scalars(db.select(Product))}
    totals = {}     # art -> summed qty
    unmatched = []
    for r in range(7, ws.max_row + 1):
        art = _art(ws.cell(r, 1).value)
        qty = _num(ws.cell(r, 5).value)
        name = ws.cell(r, 3).value
        if qty is None:
            continue
        if art and art in by_art:
            totals[art] = totals.get(art, 0) + qty
        elif art or name:
            unmatched.append((art or "", str(name or "")))
    n = 0
    for art, qty in totals.items():
        p = by_art[art]
        delta = qty - (p.stock_on_hand or 0)
        if delta:
            stock_svc.apply_movement(p, delta, "receipt", note="Coldroom opening stock")
        else:
            p.stock_on_hand = qty
        n += 1
    return n, unmatched


def _reset_store(name, kind, order):
    s = db.session.scalar(db.select(Store).filter_by(name=name))
    if s is None:
        s = Store(name=name, kind=kind, sort_order=order)
        db.session.add(s)
        db.session.flush()
    else:
        db.session.query(StoreItem).filter_by(store_id=s.id).delete(synchronize_session=False)
    return s


def import_dry(ws):
    s = _reset_store("Dry Store", "materials", 1)
    n = 0
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        qty = _num(ws.cell(r, 6).value)
        db.session.add(StoreItem(store_id=s.id, name=str(name).strip(),
                                 category=ws.cell(r, 2).value, pack_size=ws.cell(r, 3).value,
                                 uom=ws.cell(r, 4).value, origin=ws.cell(r, 5).value,
                                 quantity=qty or 0))
        n += 1
    return n


def import_two_col(sheet_name, store_name, kind, order, header_row, uom=None):
    ws = WB[sheet_name]
    s = _reset_store(store_name, kind, order)
    n = 0
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        qty = _num(ws.cell(r, 2).value)
        db.session.add(StoreItem(store_id=s.id, name=str(name).strip(),
                                 uom=uom, quantity=qty or 0))
        n += 1
    return n


WB = None


def run():
    global WB
    WB = load_workbook(FILE, data_only=True)
    stock_svc.ensure_stores()
    cold_n, unmatched = import_coldroom(WB["Coldroom Stock"])
    dry_n = import_dry(WB["Dry Store"])
    blast_n = import_two_col("BLAST FREEZER", "Blast Freezer", "production", 2, 1)
    mini_n = import_two_col("MINI BLAST STORE", "Mini Blast Store", "production", 3, 1)
    carc_n = import_two_col("CARCUS CHILLER", "Carcus Chiller", "production", 4, 2, uom="kg")
    db.session.commit()
    print(f"Coldroom -> {cold_n} products updated ({len(unmatched)} unmatched lines)")
    print(f"Dry Store: {dry_n} | Blast Freezer: {blast_n} | Mini Blast: {mini_n} | Carcus Chiller: {carc_n}")
    if unmatched:
        print("\nFirst 25 unmatched Coldroom lines (no catalogue article):")
        for art, name in unmatched[:25]:
            print(f"   {art or '(no art)'}  {name}")


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        run()
