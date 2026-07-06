"""Load the Month-over-Month product pivot into sales_history at monthly grain.

Each year sheet has customer (indent 5) -> product (indent 10) rows, with a pair
of columns (Untaxed Total, Product Quantity) per month. We flatten to one row per
customer/product/year/month. Replaces the annual pivot load (year-based reports
keep working since they aggregate by year).

Run: DATABASE_URL=sqlite:////tmp/work.db python import_sales_history_monthly.py "<xlsx>"
"""
import re
import sys
import datetime
import collections

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Customer, SalesHistory


def _norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\b(LIMITED|LTD|UGANDA|U|CO|COMPANY|ENTERPRISES|ENT|AND)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _nlead(s):
    return len(s) - len(s.lstrip(" "))


def _month(lbl):
    if not lbl:
        return None
    try:
        return datetime.datetime.strptime(str(lbl).strip(), "%B %Y")
    except ValueError:
        return None


def build_matcher():
    existing = [(c.id, c.name) for c in db.session.scalars(db.select(Customer))]
    exact = {nm: cid for cid, nm in existing}
    by_norm = collections.defaultdict(list)
    for cid, nm in existing:
        by_norm[_norm(nm)].append(cid)
    norm_list = [(_norm(nm), cid) for cid, nm in existing]

    def match(raw):
        if raw in exact:
            return exact[raw]
        n = _norm(raw)
        if n in by_norm and len(by_norm[n]) == 1:
            return by_norm[n][0]
        for en, cid in norm_list:
            if en and len(en) > 4 and (en in n or n in en):
                return cid
        return None
    return match


def run(path):
    wb = load_workbook(path, data_only=True)
    match = build_matcher()
    db.session.query(SalesHistory).delete()
    db.session.flush()
    cache = {}
    n = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        grid = list(ws.iter_rows(values_only=True))
        month_row = grid[1]
        mcols = []                       # (untaxed_col_index, year, month)
        for ci, val in enumerate(month_row):
            dt = _month(val)
            if dt:
                mcols.append((ci, dt.year, dt.month))
        cust = None
        for r in grid[3:]:
            a = r[0]
            if a is None:
                continue
            a = str(a)
            lead = _nlead(a)
            name = a.strip()
            if lead == 0 and name == "Total":
                continue
            if lead == 5:
                cust = name
            elif lead == 10 and cust is not None:
                if cust not in cache:
                    cache[cust] = match(cust)
                cid = cache[cust]
                for ci, y, m in mcols:
                    u = r[ci]
                    q = r[ci + 1] if ci + 1 < len(r) else None
                    try:
                        u = float(u)
                    except (TypeError, ValueError):
                        u = 0.0
                    if u == 0:
                        continue
                    try:
                        q = float(q)
                    except (TypeError, ValueError):
                        q = 0.0
                    db.session.add(SalesHistory(
                        customer_id=cid, customer_name=cust, product=name,
                        year=y, month=m, revenue=round(u, 2), quantity=q,
                        is_return=(u < 0)))
                    n += 1
        db.session.flush()
    db.session.commit()
    matched = sum(1 for v in cache.values() if v)
    print(f"Loaded {n} monthly rows. Customers {len(cache)} ({matched} matched).")
    for y in (2024, 2025, 2026):
        tot = sum(float(s.revenue) for s in db.session.scalars(
            db.select(SalesHistory).filter_by(year=y)))
        print(f"  {y}: {tot:,.2f} UGX")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else \
        "imports/Month-over-Month Pivot Invoices Analysis 2024, 2025, 2026.xlsx"
    app = create_app()
    with app.app_context():
        run(path)
