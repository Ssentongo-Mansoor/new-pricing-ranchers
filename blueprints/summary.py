"""Dashboard / summary table with category filter, margin colouring, and exports."""
import io

from flask import (Blueprint, render_template, request, send_file)
from flask_login import login_required, current_user

from services.costing_models import Recipe, Category, Setting
from services.costing_engine import recipe_cost_per_kg, pricing_for_cost
from services.costing_auth import fmt_money

summary_bp = Blueprint("summary", __name__, url_prefix="/costing/summary")


def _build_rows(show_all=False, cat_id=None):
    query = Recipe.query
    if not show_all:
        query = query.filter_by(status="active")
    if cat_id:
        query = query.filter_by(category_id=cat_id)
    recipes = query.all()
    wm = Setting.get_float("wholesale_margin", 0.47)
    rm = Setting.get_float("rrp_margin", 0.15)
    vat = Setting.get_float("vat_rate", 0.18)
    rows = []
    for r in recipes:
        cost = recipe_cost_per_kg(r)
        pr = pricing_for_cost(cost, wm, rm, vat)
        rows.append({
            "recipe": r,
            "name": r.name,
            "category": r.category.name if r.category else "—",
            "category_order": r.category.display_order if r.category else 999,
            "status": r.status,
            "cost": cost,
            "wholesale_excl": pr["wholesale_excl"],
            "wholesale_incl": pr["wholesale_incl"],
            "rrp_excl": pr["rrp_excl"],
            "rrp_incl": pr["rrp_incl"],
            "margin_pct": pr["margin_pct"],
        })
    rows.sort(key=lambda x: (x["category_order"], x["name"]))
    return rows


@summary_bp.route("/")
@login_required
def dashboard():
    show_all = request.args.get("show") == "all"
    cat_id = request.args.get("category", type=int)
    rows = _build_rows(show_all, cat_id)
    categories = Category.query.order_by(Category.display_order).all()
    low = Setting.get_float("margin_threshold_low", 25)
    high = Setting.get_float("margin_threshold_high", 40)
    active_count = Recipe.query.filter_by(status="active").count()
    total_count = Recipe.query.count()
    below = sum(1 for r in rows if r["margin_pct"] < low)
    return render_template("costing/dashboard.html", rows=rows, categories=categories,
                           show_all=show_all, cat_id=cat_id, low=low, high=high,
                           active_count=active_count, total_count=total_count,
                           below=below, hide_costs=False)


@summary_bp.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    rows = _build_rows(request.args.get("show") == "all",
                       request.args.get("category", type=int))
    low = Setting.get_float("margin_threshold_low", 25)
    high = Setting.get_float("margin_threshold_high", 40)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Summary"
    headers = ["Product", "Category", "Status", "Cost/kg", "Wholesale Excl VAT",
               "Wholesale Incl VAT", "RRP Excl VAT", "RRP Incl VAT", "Margin %"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="8B1A1A")
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    for r in rows:
        ws.append([r["name"], r["category"], r["status"], round(r["cost"]),
                   round(r["wholesale_excl"]), round(r["wholesale_incl"]),
                   round(r["rrp_excl"]), round(r["rrp_incl"]),
                   round(r["margin_pct"], 1)])
        cell = ws.cell(ws.max_row, 9)
        cell.fill = green if r["margin_pct"] >= high else (red if r["margin_pct"] < low else yellow)
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["A"].width = 36
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="product_summary.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@summary_bp.route("/export/pdf")
@login_required
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet
    rows = _build_rows(request.args.get("show") == "all",
                       request.args.get("category", type=int))
    low = Setting.get_float("margin_threshold_low", 25)
    high = Setting.get_float("margin_threshold_high", 40)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Ranchers Finest — Product Price List", styles["Title"]),
             Spacer(1, 6)]
    data = [["Product", "Category", "Cost/kg", "Wholesale\nExcl VAT",
             "Wholesale\nIncl VAT", "RRP\nExcl VAT", "RRP\nIncl VAT", "Margin %"]]
    for r in rows:
        data.append([r["name"][:34], r["category"][:22], fmt_money(r["cost"]),
                     fmt_money(r["wholesale_excl"]), fmt_money(r["wholesale_incl"]),
                     fmt_money(r["rrp_excl"]), fmt_money(r["rrp_incl"]),
                     f"{r['margin_pct']:.0f}%"])
    table = Table(data, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B1A1A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f0ef")]),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
    ])
    for i, r in enumerate(rows, start=1):
        if r["margin_pct"] >= high:
            c = colors.HexColor("#C6EFCE")
        elif r["margin_pct"] < low:
            c = colors.HexColor("#FFC7CE")
        else:
            c = colors.HexColor("#FFEB9C")
        style.add("BACKGROUND", (7, i), (7, i), c)
    table.setStyle(style)
    elems.append(table)
    doc.build(elems)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="price_list.pdf",
                     mimetype="application/pdf")


@summary_bp.before_request
def _costing_gate():
    from flask_login import current_user
    from flask import abort
    if not current_user.is_authenticated:
        abort(401)
    from services.costing_auth import require_costing_view
    require_costing_view()
